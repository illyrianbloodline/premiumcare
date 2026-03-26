"""
MedPlatform – Main Flask Application
Run: python app.py
Access: http://localhost:5000
Default login: admin@practice.local / admin123
"""

import os, base64, json
from datetime import date, timedelta, datetime
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash
try:
    from flask_talisman import Talisman
    _talisman_available = True
except ImportError:
    _talisman_available = False

try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address
    _limiter_available = True
except ImportError:
    _limiter_available = False
from flask_bcrypt import Bcrypt
from functools import wraps
import database as db
from groq import Groq as _GroqClient
from config import SECRET_KEY, GROQ_API_KEY, GEMINI_API_KEY, GEMINI_MODEL, UPLOAD_FOLDER, MAX_UPLOAD_MB, PRACTICE_NAME, SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASSWORD, SMTP_FROM
from werkzeug.utils import secure_filename
from flask import send_from_directory
import security as sec
from ai_firewall import AIFirewall as _AIFirewall
from security import sanitise_form_data, sanitise_medical_notes, sanitise_plain_input

# ── Gemini ia helper ─────────────────────────────────────────────────────────
import os as _os

# Fallback model chain — tries each in order until one works
GEMINI_FALLBACK_MODELS = [
    GEMINI_MODEL,           # primary from config.py
    "gemini-2.0-flash",
    "gemini-2.0-flash-lite",
    "gemini-2.5-flash-lite",
]

def _gemini(prompt, image_data=None, media_type=None):
    """
    Text queries  → Groq / Llama 3.3 70B  (free, 14,400 req/day)
    Image queries → Gemini with automatic fallback between models.
    If one model hits quota or is unavailable, tries the next one.
    """
    if image_data and media_type:
        # ── Image path: use Gemini with fallback ───────────────
        import google.generativeai as _genai
        import base64 as _b64
        gemini_key = _os.environ.get("GEMINI_API_KEY", GEMINI_API_KEY)
        if not gemini_key or gemini_key in ("paste-your-gemini-key-here", ""):
            return "Scan analysis requires a Gemini API key. Add GEMINI_API_KEY to config.py."
        _genai.configure(api_key=gemini_key)
        raw = _b64.b64decode(image_data) if isinstance(image_data, str) else image_data
        last_error = None
        for model_name in GEMINI_FALLBACK_MODELS:
            try:
                model = _genai.GenerativeModel(
                    model_name=model_name,
                    system_instruction=(
                        "You are a helpful assistant that describes the contents of images, "
                        "including medical scans, diagrams, and radiological images for "
                        "educational and informational purposes. Describe all visible "
                        "anatomical structures and landmarks objectively. Never refuse to "
                        "describe an image."
                    )
                )
                response = model.generate_content([
                    {"mime_type": media_type, "data": raw},
                    prompt
                ])
                return response.text
            except Exception as e:
                last_error = e
                err_str = str(e)
                # Only fall through on quota/not-found errors
                if any(code in err_str for code in ['429', '404', 'NOT_FOUND', 'RESOURCE_EXHAUSTED']):
                    continue
                raise  # re-raise other errors immediately
        raise last_error
    else:
        # ── Text path: use Groq ────────────────────────────────
        client = _GroqClient(api_key=_os.environ.get("GROQ_API_KEY", GROQ_API_KEY))
        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=1500,
            temperature=0.3,
        )
        return response.choices[0].message.content


app = Flask(__name__)
# ── Guard: refuse to start with a default or weak secret key ─────────────────
_UNSAFE_KEYS = {
    "change-this-to-a-long-random-string-in-production",
    "change-this", "secret", "dev", "development", "test", "",
}
if SECRET_KEY.lower() in _UNSAFE_KEYS or len(SECRET_KEY) < 32:
    raise RuntimeError(
        "\n\n[SECURITY] SECRET_KEY is unsafe or too short.\n"
        "Set a strong random key in config.py or via the SECRET_KEY env var:\n"
        "  python -c \"import secrets; print(secrets.token_hex(32))\"\n"
    )
app.secret_key = SECRET_KEY
# ── Secure session cookie flags ────────────────────────────────────────────────
_HTTPS_ACTIVE = os.environ.get('FORCE_HTTPS', 'false').lower() == 'true'
app.config['SESSION_COOKIE_HTTPONLY']  = True
app.config['SESSION_COOKIE_SAMESITE']  = 'Lax'
app.config['SESSION_COOKIE_SECURE']    = _HTTPS_ACTIVE
app.config['SESSION_COOKIE_NAME']      = '__Host-medplat_sess' if _HTTPS_ACTIVE else 'medplat_sess'
app.config['PERMANENT_SESSION_LIFETIME'] = 3600
app.config['WTF_CSRF_ENABLED']         = True
app.config['WTF_CSRF_TIME_LIMIT']      = 3600

# ── Talisman — HTTP security headers (safe mode) ──────────────
# CSP disabled to avoid breaking inline styles/scripts in templates.
# All other headers active: X-Frame-Options, X-Content-Type-Options,
# Referrer-Policy. Enable force_https=True once you have SSL.
if _talisman_available:
    Talisman(
        app,
        force_https=False,
        strict_transport_security=False,
        session_cookie_secure=False,
        content_security_policy=False,
        frame_options='DENY',
        referrer_policy='strict-origin-when-cross-origin',
    )

# ── Flask-Limiter ─────────────────────────────────────────────
# Redis-backed rate limiting. Falls back to memory if Redis unavailable.
if _limiter_available:
    _redis_url = os.environ.get('REDIS_URL', 'redis://localhost:6379/0')
    try:
        import redis as _r; _r.Redis.from_url(_redis_url, socket_connect_timeout=1).ping()
        _limiter_storage = f"redis://{_redis_url.split('://',1)[-1]}"
    except Exception:
        _limiter_storage = "memory://"
    limiter = Limiter(
        app=app,
        key_func=get_remote_address,
        storage_uri=_limiter_storage,
        default_limits=["200 per hour"],
        headers_enabled=True,
    )
else:
    limiter = None

app.config['MAX_CONTENT_LENGTH'] = MAX_UPLOAD_MB * 1024 * 1024
bcrypt = Bcrypt(app)

# ── CSRF protection (flask-wtf) ───────────────────────────────────────────────
try:
    from flask_wtf.csrf import CSRFProtect, CSRFError
    csrf = CSRFProtect(app)

    @app.errorhandler(CSRFError)
    def csrf_error(e):
        import logging; logging.getLogger('security').warning('CSRF token missing/invalid: %s', e.description)
        if request.is_json:
            from flask import jsonify
            return jsonify({'ok': False, 'error': 'CSRF token invalid'}), 400
        flash('Sesioni skadoi. Ju lutem provoni përsëri.', 'danger')
        return redirect(url_for('login'))
except ImportError:
    csrf = None

# ── AI Firewall (Chinese Wall) — strips PHI before every AI call ──────────────
_fw = _AIFirewall(strict=True, enable_canaries=True)


# ── Flask-Limiter — rate limiting ─────────────────────────────
try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address
    limiter = Limiter(
        app=app,
        key_func=get_remote_address,
        default_limits=["200 per minute", "2000 per hour"],
        storage_uri=os.getenv("REDIS_URL", "memory://"),
    )
    _limiter_available = True
except ImportError:
    limiter = None
    _limiter_available = False
    print("[WARNING] flask-limiter not installed. Run: pip install flask-limiter")

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'tiff', 'dcm', 'pdf'}

# ── Security middleware ────────────────────────────────────────────────────────

@app.before_request
def before_request():
    # Block known attack paths
    path = request.path.lower()
    bad_paths = ['/wp-admin','/.env','/phpmyadmin','/shell','/cmd','/config.php']
    if any(p in path for p in bad_paths):
        try: db.log_activity(None, '🚨 ATTACK PATH BLOCKED', f"IP:{request.remote_addr} Path:{request.path}")
        except: pass
        abort(404)
    # Block SQL injection in query string
    qs = (request.query_string or b'').decode('utf-8','ignore').lower()
    sql_signs = ["' --","union select","drop table","sleep(","1=1--"]
    if any(s in qs for s in sql_signs):
        try: db.log_activity(None, '🚨 SQL INJECTION BLOCKED', f"IP:{request.remote_addr} QS:{qs[:100]}")
        except: pass
        abort(400)
    # Check session timeout
    if sec.check_session_timeout():
        flash('Session expired. Please log in again.', 'warning')
        return redirect(url_for('login'))
    # Detect session hijacking (IP mismatch)
    if 'staff_id' in session:
        bound = session.get('bound_ip')
        current = request.remote_addr
        if bound and bound != current:
            try: db.log_activity(session.get('staff_id'), '🚨 SESSION HIJACK DETECTED',
                    f"Session moved from {bound} to {current}")
            except: pass
            session.clear()
            flash('Security alert: session invalidated.', 'danger')
            return redirect(url_for('login'))@app.after_request
def add_security_headers(response):
    return sec.apply_security_headers(response)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ── Permission definitions ─────────────────────────────────────────────────────

ALL_PERMISSIONS = [
    # (key, section, label, description)
    ('patients_view',  'Patients',    'View patients',        'See patient list and demographics'),
    ('patients_edit',  'Patients',    'Edit patients',        'Create, edit and delete patient records'),
    ('clinical',       'Clinical',    'Clinical notes & AI',  'Visit notes, diagnoses, AI tools'),
    ('appointments',   'Operations',  'Appointments',         'View and manage the calendar'),
    ('billing',        'Finance',     'Billing & invoices',   'Invoices, payments, financial reports'),
    ('team_view',      'HR',          'View staff',           'See the team directory'),
    ('team_edit',      'HR',          'Manage staff',         'Add, edit and deactivate staff members'),
    ('activity_log',   'Admin',       'Activity log',         'View the full audit trail'),
]

ROLE_PRESETS = {
    'admin':        ['patients_view','patients_edit','clinical','appointments','billing','team_view','team_edit','activity_log'],
    'doctor':       ['patients_view','patients_edit','clinical','appointments'],
    'nurse':        ['patients_view','clinical','appointments'],
    'receptionist': ['patients_view','appointments','billing'],
    'finance':      ['billing'],
    'hr':           ['team_view','team_edit'],
    'custom':       [],
}

def parse_permissions(perm_json):
    """Return list of permission keys from stored JSON string, or [] on error."""
    if not perm_json:
        return []
    try:
        return json.loads(perm_json)
    except Exception:
        return []

def build_permissions_json(form_data):
    """Read checkbox values from form and return JSON string."""
    granted = [key for (key, *_) in ALL_PERMISSIONS if form_data.get(f'perm_{key}')]
    return json.dumps(granted)

def has_perm(perm):
    """Check if current session user has a permission. Admin always returns True."""
    if session.get('role') == 'admin':
        return True
    return perm in session.get('permissions', [])

# Make has_perm available in every Jinja template
@app.context_processor
def inject_permissions():
    return dict(has_perm=has_perm, ALL_PERMISSIONS=ALL_PERMISSIONS, ROLE_PRESETS=ROLE_PRESETS)

@app.context_processor
def inject_translations():
    """Inject English translations as `t` into every template."""
    from lang import get_t
    lang = session.get('lang', 'sq')  # Albanian default
    t = get_t(lang)
    # extra keys used by uploaded templates
    t.setdefault('back', 'Back')
    t.setdefault('save', 'Save')
    t.setdefault('cancel', 'Cancel')
    t.setdefault('confirm_delete', 'Are you sure you want to delete this record?')
    t.setdefault('id', 'ID')
    t.setdefault('status', 'Status')
    t.setdefault('appointment_type', 'Type')
    # Patient form keys missing from lang.py
    t.setdefault('pf_medical',     'Medical History')
    t.setdefault('pf_history',     'Conditions / History')
    t.setdefault('pf_medications', 'Medications')
    t.setdefault('gender',         'Gender')
    t.setdefault('new_patient',    'New Patient')
    unread = db.get_unread_count(session['staff_id']) if 'staff_id' in session else 0
    pending_tasks = db.get_pending_task_counts(session['staff_id']) if 'staff_id' in session else 0
    return dict(t=t, unread_count=unread, pending_tasks=pending_tasks)

def _lang_instruction():
    """Returns a language instruction string based on the current session language."""
    lang = session.get('lang', 'sq')
    if lang == 'sq':
        return "IMPORTANT: You MUST respond entirely in Albanian (Shqip) language. Do not use English in your response."
    return "Respond in English."

def _ai_suffix():
    """Structure instructions — no disclaimer (already shown in UI)."""
    lang = session.get('lang', 'sq')
    if lang == 'sq':
        return """

---
RREGULL: MOS shto asnjë paralajmërim ose disclaimer — mjeku e di tashmë. Mbyll me:

**📋 Përmbledhje:**
• [Pika 1]
• [Pika 2]
• [Pika 3]

**🩺 Konkluzion:** [1-2 fjali]
"""
    else:
        return """

---
RULE: Do NOT add any disclaimer or warning — physician already knows. Close with:

**📋 Summary:**
• [Point 1]
• [Point 2]
• [Point 3]

**🩺 Conclusion:** [1-2 sentences]
"""


# ── Auth helpers ───────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'staff_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

@app.context_processor
def inject_security():
    """Inject CSP nonce and security helpers into every template."""
    from security import get_csp_nonce
    return dict(csp_nonce=get_csp_nonce())

@app.before_request
def enforce_json_content_type():
    """Reject POST requests to /api/* that don't send Content-Type: application/json."""
    if (request.path.startswith('/api/')
            and request.method == 'POST'
            and request.content_length
            and request.content_length > 0
            and not request.is_json):
        return jsonify({'ok': False, 'error': 'Content-Type must be application/json'}), 415


@app.context_processor
def inject_settings():
    """Inject practice settings and security helpers into every template."""
    try:
        settings = db.get_settings()
    except Exception:
        settings = {}
    # Expose CSP nonce to templates: <script nonce="{{ csp_nonce() }}">
    return dict(
        practice_settings=settings,
        csp_nonce=sec.get_csp_nonce,
    )

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'admin':
            flash('Admin access required.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return login_required(decorated)

def permission_required(perm):
    """Decorator: requires a specific permission. Admin always passes."""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not has_perm(perm):
                flash('You do not have permission to access that page.', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return login_required(decorated)
    return decorator

def current_user():
    return {
        'id':          session.get('staff_id'),
        'name':        session.get('staff_name'),
        'role':        session.get('role'),
        'permissions': session.get('permissions', []),
    }

# Magic bytes for allowed upload formats (first 8 bytes)
_MAGIC_BYTES = {
    b'\x89PNG':     'png',
    b'\xff\xd8\xff': 'jpg',
    b'GIF8':        'gif',
    b'BM':          'bmp',
    b'II*\x00':    'tiff',
    b'MM\x00*':    'tiff',
    b'%PDF':        'pdf',
}

def allowed_file(filename: str) -> bool:
    """Extension check only — use validate_upload() for full magic-byte check."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def validate_upload(file_storage) -> tuple[bool, str]:
    """
    Full upload validation: extension check + magic-byte verification.
    Prevents content-type spoofing (e.g. .jpg file containing HTML/JS).

    Returns (ok: bool, error_message: str).
    """
    filename = file_storage.filename or ''
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    if ext not in ALLOWED_EXTENSIONS:
        return False, f"File type .{ext} not allowed"

    # Read first 8 bytes for magic-byte check (then seek back)
    header = file_storage.read(8)
    file_storage.seek(0)

    # PDF and DICOM — check magic bytes
    for magic, fmt in _MAGIC_BYTES.items():
        if header.startswith(magic):
            if fmt in ('jpg','jpeg') and ext in ('jpg','jpeg'):
                return True, ""
            if fmt == ext or (fmt == 'tiff' and ext in ('tiff','tif')):
                return True, ""
            if fmt == 'pdf' and ext == 'pdf':
                return True, ""
            if fmt == 'png' and ext == 'png':
                return True, ""

    # DICOM has no universal magic byte — trust extension
    if ext == 'dcm':
        return True, ""

    # GIF/BMP/TIFF — extension-only check (varied magic bytes)
    if ext in ('gif', 'bmp', 'tiff', 'tif'):
        return True, ""

    # JPEG magic found but extension mismatch — still allow
    if header[:3] == b'\xff\xd8\xff' and ext in ('jpg','jpeg'):
        return True, ""

    return False, f"File content does not match extension .{ext}"

# ── Auth ───────────────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    ip = request.remote_addr or '0.0.0.0'
    locked, minutes = sec.is_locked_out(ip)
    if locked:
        return render_template('login.html', practice_name=PRACTICE_NAME,
                               locked=True, minutes=minutes)

    if request.method == 'POST':
        email    = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        staff    = db.get_staff_by_email(email)

        if staff and bcrypt.check_password_hash(staff['password'], password):
            if not staff.get('active', True):
                flash('Account is deactivated. Contact your administrator.', 'danger')
                return render_template('login.html', practice_name=PRACTICE_NAME)

            # ── Canary / honeypot check ─────────────────────────────────────
            CANARY_EMAILS = ['backup.admin@practice.local',
                             'admin.backup@practice.local',
                             'sysadmin@practice.local']
            if email in CANARY_EMAILS:
                db.log_activity(None, '🚨 CANARY TRIGGERED — POSSIBLE BREACH',
                    f"Canary email: {email} | IP: {ip} | Time: {datetime.now()}")
                flash('Invalid email or password.', 'danger')
                return render_template('login.html', practice_name=PRACTICE_NAME,
                                       t=__import__('lang').get_t(session.get('lang','sq')))

            # Credentials valid — set up session
            sec.clear_login_attempts(ip)
            session['bound_ip'] = ip   # bind session to this IP
            session.clear()
            session.regenerate() if hasattr(session, 'regenerate') else None
            session['staff_id']       = staff['id']
            session['staff_name']     = staff['name']
            session['role']           = staff['role']
            session['last_activity']  = datetime.now().isoformat()
            session['login_ip']       = ip
            session['lang']           = request.form.get('lang', 'sq')

            if staff['role'] == 'admin':
                session['permissions'] = [key for (key, *_) in ALL_PERMISSIONS]
            else:
                session['permissions'] = parse_permissions(staff.get('permissions'))

            # Check if MFA is configured for this account
            if staff.get('mfa_secret'):
                session['mfa_pending_id'] = staff['id']
                session.pop('mfa_verified', None)
                return redirect(url_for('mfa_verify'))
            else:
                session['mfa_verified'] = True   # MFA not yet enrolled — skip for now

            db.log_activity(staff['id'], 'Login', f"Successful login from {ip}")
            return redirect(url_for('dashboard'))

        # Failed login
        locked, lockout_mins = sec.record_failed_login(ip, email)
        remaining = sec.get_remaining_attempts(ip)
        # Per-email daily limit tracked in Redis by Flask-Limiter (set on route)
        db.log_activity(None, 'SECURITY — Failed login', f"Email: {email} | IP: {ip}")

        if locked:
            return render_template('login.html', practice_name=PRACTICE_NAME,
                                   locked=True, minutes=lockout_mins)
        flash(f'Invalid email or password. {remaining} attempt(s) remaining before lockout.', 'danger')

    attempts_left = sec.get_remaining_attempts(ip)
    return render_template('login.html', practice_name=PRACTICE_NAME,
                           attempts_left=attempts_left)

@app.route('/logout')
def logout():
    if 'staff_id' in session:
        db.log_activity(session['staff_id'], 'Logout', f"{session['staff_name']} logged out")
    session.clear()
    return redirect(url_for('login'))

# ── Dashboard ──────────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def dashboard():
    import json as _json
    perms    = session.get('permissions', [])
    is_admin = (session.get('role') == 'admin')
    def p(key): return is_admin or key in perms

    ctx = dict(user=current_user(), practice_name=PRACTICE_NAME, now=datetime.now())

    # ── Clinical ──────────────────────────────────────────────
    if p('patients_view'):
        pts = db.get_all_patients()
        ctx['total_patients']  = len(pts)
        ctx['recent_patients'] = db.get_recent_patients(6)
    if p('appointments'):
        today_appts = db.get_appointments_for_day(date.today().isoformat())
        waiting     = db.get_waiting_room()
        ctx['today_appts']      = len(today_appts)
        ctx['today_appts_list'] = today_appts[:7]
        ctx['waiting_count']    = len(waiting)
        ctx['waiting_list']     = waiting[:8]
    if p('clinical'):
        ctx['ai_stats'] = db.get_ai_usage_stats()

    # ── Finance ───────────────────────────────────────────────
    if p('billing'):
        all_inv  = db.get_invoices()
        unpaid   = [i for i in all_inv if i['status'] != 'paid']
        overdue  = [i for i in unpaid if i.get('due_date') and i['due_date'] < date.today()]
        paid_amt = sum(float(i['amount_paid'] or 0) for i in all_inv)
        due_amt  = sum(float(i['total_amount'] or 0) for i in unpaid)
        total_billed = sum(float(i['total_amount'] or 0) for i in all_inv)

        monthly  = db.get_monthly_revenue(6)
        ctx['finance_total_invoices']  = len(all_inv)
        ctx['finance_unpaid_count']    = len(unpaid)
        ctx['finance_overdue_count']   = len(overdue)
        ctx['finance_paid_amount']     = paid_amt
        ctx['finance_due_amount']      = due_amt
        ctx['finance_total_billed']    = total_billed
        ctx['finance_recent']          = all_inv[:8]
        ctx['finance_chart_labels']    = _json.dumps([r['month_label'] for r in monthly])
        ctx['finance_chart_billed']    = _json.dumps([float(r['billed']) for r in monthly])
        ctx['finance_chart_collected'] = _json.dumps([float(r['collected']) for r in monthly])

    # ── HR ────────────────────────────────────────────────────
    if p('team_view'):
        staff = db.get_all_staff()
        roles = db.get_staff_role_distribution()
        active   = [s for s in staff if s['active']]
        inactive = [s for s in staff if not s['active']]
        ctx['total_staff']      = len(active)
        ctx['inactive_staff']   = len(inactive)
        ctx['staff_list']       = active
        ctx['staff_roles']      = roles
        ctx['staff_role_labels']  = _json.dumps([r['role'].title() for r in roles])
        ctx['staff_role_counts']  = _json.dumps([r['count'] for r in roles])

    # ── Operations ────────────────────────────────────────────
    if p('activity_log'):
        ctx['logs'] = db.get_activity_log(12)

    # Inbox (everyone)
    ctx['unread_messages'] = db.get_unread_count(session['staff_id'])

    # ── Smart Today Alerts ────────────────────────────────────────────
    try:
        ctx['smart_alerts'] = db.get_smart_today_alerts()
    except Exception as _sa_err:
        import traceback
        print('[SMART ALERTS ERROR]', traceback.format_exc())
        ctx['smart_alerts'] = []

    return render_template('dashboard.html', **ctx)

# ── Patients ───────────────────────────────────────────────────────────────────

@app.route('/patients')
@permission_required('patients_view')
@sec.audit_bulk_access('patient_list')
def patients():
    search = request.args.get('q', '')
    rows   = db.get_all_patients(search)
    return render_template('patients.html', user=current_user(),
                           practice_name=PRACTICE_NAME, patients=rows, search=search,
                           today=date.today())

@app.route('/patients/new', methods=['GET', 'POST'])
@permission_required('patients_edit')
def patient_new():
    if request.method == 'POST':
        data = {k: v.strip() or None for k, v in request.form.items()}
        pid  = db.save_patient(data, staff_id=session['staff_id'])
        db.log_activity(session['staff_id'], 'New patient',
                        f"{data.get('first_name')} {data.get('last_name')}")
        flash('Patient record created.', 'success')
        return redirect(url_for('patient_view', pid=pid))
    return render_template('patient_form.html', user=current_user(),
                           practice_name=PRACTICE_NAME, patient=None)

@app.route('/patients/<int:pid>')
@permission_required('patients_view')
@sec.audit_access('patient_record', id_arg='pid')
def patient_view(pid):
    patient = db.get_patient(pid)
    if not patient:
        flash('Patient not found.', 'danger')
        return redirect(url_for('patients'))
    analyses     = db.get_symptom_analyses(pid)
    scans        = db.get_scan_analyses(pid)
    visits       = db.get_visit_notes_for_patient(pid)
    appointments = db.get_appointments_for_patient(pid)
    staff_list   = db.get_all_staff()
    vaccinations  = db.get_vaccinations(pid)      if hasattr(db, 'get_vaccinations')       else []
    recalls       = db.get_recalls(pid)            if hasattr(db, 'get_recalls')            else []
    documents     = db.get_patient_documents(pid)  if hasattr(db, 'get_patient_documents')  else []
    lab_tests     = db.get_lab_tests(pid)          if hasattr(db, 'get_lab_tests')          else []
    prescriptions = db.get_prescriptions(pid)      if hasattr(db, 'get_prescriptions')      else []
    invoices      = db.get_invoices(pid)           if hasattr(db, 'get_invoices')           else []
    return render_template('patient_view.html', user=current_user(),
                           practice_name=PRACTICE_NAME, patient=patient,
                           analyses=analyses, scans=scans,
                           visits=visits, appointments=appointments,
                           staff_list=staff_list,
                           vaccinations=vaccinations,
                           recalls=recalls,
                           documents=documents,
                           lab_tests=lab_tests,
                           prescriptions=prescriptions,
                           invoices=invoices,
                           today=date.today())

@app.route('/patients/<int:pid>/edit', methods=['GET', 'POST'])
@permission_required('patients_edit')
@sec.audit_access('patient_record', id_arg='pid', action='EDIT')
def patient_edit(pid):
    patient = db.get_patient(pid)
    if not patient:
        return redirect(url_for('patients'))
    sec.log_phi_access(session['staff_id'], 'patient_record', pid, 'EDIT')
    if request.method == 'POST':
        data = sanitise_form_data(request.form.to_dict())
        db.save_patient(data, pid=pid)
        db.log_activity(session['staff_id'], 'Updated patient',
                        f"{patient['first_name']} {patient['last_name']}")
        flash('Patient record updated.', 'success')
        return redirect(url_for('patient_view', pid=pid))
    return render_template('patient_form.html', user=current_user(),
                           practice_name=PRACTICE_NAME, patient=patient)

@app.route('/patients/<int:pid>/delete', methods=['POST'])
@permission_required('patients_edit')
@sec.audit_access('patient_record', id_arg='pid', action='DELETE')
def patient_delete(pid):
    p = db.get_patient(pid)
    if p:
        db.delete_patient(pid)
        db.log_activity(session['staff_id'], 'Deleted patient',
                        f"{p['first_name']} {p['last_name']}")
        flash('Patient record deleted.', 'success')
    return redirect(url_for('patients'))

# ── Patient Bulk Import ──────────────────────────────────────────────────────

@app.route('/patients/import-template')
@login_required
def patient_import_template():
    """Download the Excel import template."""
    # Try multiple paths to find the file
    candidates = [
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'patient_import_template.xlsx'),
        os.path.join(os.getcwd(), 'patient_import_template.xlsx'),
        os.path.join(app.root_path, 'patient_import_template.xlsx'),
    ]
    template_path = None
    for path in candidates:
        if os.path.isfile(path):
            template_path = path
            break
    if not template_path:
        flash(f'Skedari nuk u gjet. Vendoseni patient_import_template.xlsx në: {candidates[0]}', 'danger')
        return redirect(url_for('patients'))
    directory = os.path.dirname(template_path)
    return send_from_directory(
        directory,
        'patient_import_template.xlsx',
        as_attachment=True,
        download_name='MedPlatform_Shablloni_Importit.xlsx',
    )


@app.route('/patients/import', methods=['POST'])
@login_required
def patient_import():
    """Process uploaded Excel file and bulk-create patients."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        flash('openpyxl not installed. Run: pip install openpyxl', 'danger')
        return redirect(url_for('patients'))

    file = request.files.get('import_file')
    if not file or not file.filename.endswith('.xlsx'):
        flash('Ju lutem ngarkoni një skedar .xlsx të vlefshëm.', 'danger')
        return redirect(url_for('patients'))

    # Column mapping: Excel column index (1-based) → patient field
    COL_MAP = {
        1:  'first_name',    2:  'last_name',      3:  'dob',
        4:  'gender',        5:  'blood_type',      6:  'insurance_nr',
        7:  'address',       8:  'phone',           9:  'email',
        10: 'emergency_contact', 11: 'family_doctor',
        12: 'medical_history',   13: 'surgeries',      14: 'family_history',
        15: 'drug_allergies',    16: 'env_allergies',  17: 'medications',
        18: 'supplements',       19: 'past_medications',
        20: 'weight',        21: 'height',         22: 'blood_pressure',
        23: 'heart_rate',    24: 'temperature',    25: 'oxygen_sat',
        26: 'blood_glucose', 27: 'cholesterol',
        28: 'icd_codes',     29: 'referrals',      30: 'notes',
    }

    try:
        wb  = load_workbook(file, data_only=True)
        ws  = wb.active
    except Exception as e:
        flash(f'Gabim duke lexuar skedarin: {e}', 'danger')
        return redirect(url_for('patients'))

    created = 0
    skipped = 0
    errors  = []

    # Data starts at row 9 (rows 1-4 = headers, rows 5-7 = country examples, row 8 = filled example)
    for row_num in range(9, ws.max_row + 1):
        row = [ws.cell(row=row_num, column=c).value for c in range(1, 31)]

        # Skip completely empty rows
        if not any(v for v in row if v is not None and str(v).strip()):
            continue

        # Skip header/example rows
        first = str(row[0] or '').strip()
        if (first.lower() in ('emri', 'emri *', 'shembull', 'example', 'rresht shembull')
                or first.startswith(('🇽🇰','🇦🇱','🇲🇰','⬇','✏'))):
            skipped += 1
            continue

        # Required fields
        first_name = str(row[0] or '').strip()
        last_name  = str(row[1] or '').strip()
        if not first_name or not last_name:
            errors.append(f'Rreshti {row_num}: Emri dhe Mbiemri janë të detyrueshme.')
            skipped += 1
            continue

        # Build patient dict
        data = {}
        for col_idx, field in COL_MAP.items():
            raw = row[col_idx - 1]
            if raw is None:
                data[field] = None
                continue
            val = str(raw).strip() if raw != '' else None

            # Parse DOB from DD/MM/YYYY
            if field == 'dob' and val:
                from datetime import datetime as _dt
                for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d.%m.%Y', '%m/%d/%Y'):
                    try:
                        data[field] = _dt.strptime(val, fmt).date().isoformat()
                        break
                    except ValueError:
                        continue
                else:
                    data[field] = None  # unparseable date
            else:
                data[field] = val

        try:
            # Ensure all patient fields have at least None so save_patient doesn't fail
            import database as _db2
            for field in _db2.PATIENT_FIELDS:
                if field not in data:
                    data[field] = None
            db.save_patient(data, staff_id=session['staff_id'])
            created += 1
        except Exception as e:
            import traceback; traceback.print_exc()
            errors.append(f'Rreshti {row_num} ({first_name} {last_name}): {str(e)[:120]}')
            skipped += 1

    db.log_activity(session['staff_id'], 'Bulk patient import',
                    f'{created} pacientë u importuan, {skipped} u kaluan')

    # Store detailed errors in session so the template can show them nicely
    session['import_result'] = {
        'created': created,
        'skipped': skipped,
        'errors': errors[:20],  # max 20 errors shown
    }

    # Result shown via popup modal — no flash needed

    return redirect(url_for('patients'))


@app.route('/patients/import-result/clear', methods=['POST'])
@login_required
def clear_import_result():
    """Clear the import result from session after modal is dismissed."""
    session.pop('import_result', None)
    return jsonify({'ok': True})


# ── Visit Notes ────────────────────────────────────────────────────────────────

@app.route('/patients/<int:pid>/visits/new', methods=['POST'])
@permission_required('clinical')
def visit_note_new(pid):
    data = {
        'patient_id':      pid,
        'doctor_id':       request.form.get('doctor_id') or None,
        'appointment_id':  request.form.get('appointment_id') or None,
        'visit_date':      request.form.get('visit_date'),
        'visit_type':      request.form.get('visit_type'),
        'status':          request.form.get('status', 'completed'),
        'chief_complaint': sanitise_medical_notes(request.form.get('chief_complaint','')),
        'diagnosis':       sanitise_medical_notes(request.form.get('diagnosis','')),
        'treatment':       sanitise_medical_notes(request.form.get('treatment','')),
        'clinical_notes':  sanitise_medical_notes(request.form.get('clinical_notes','')),
        'follow_up':       sanitise_medical_notes(request.form.get('follow_up','')),
    }
    if not data['visit_date']:
        flash('Visit date is required.', 'danger')
        return redirect(url_for('patient_view', pid=pid))
    db.save_visit_note(data, created_by=session['staff_id'])
    db.log_activity(session['staff_id'], 'New visit note',
                    f"Patient ID {pid} - {data['visit_date']}")
    flash('Visit note saved successfully.', 'success')
    return redirect(url_for('patient_view', pid=pid) + '#tab-visits')

@app.route('/patients/<int:pid>/visits/<int:note_id>/edit', methods=['POST'])
@permission_required('clinical')
def visit_note_edit(pid, note_id):
    note = db.get_visit_note(note_id)
    if not note or note['patient_id'] != pid:
        flash('Visit note not found.', 'danger')
        return redirect(url_for('patient_view', pid=pid))
    data = {
        'doctor_id':       request.form.get('doctor_id') or note['doctor_id'],
        'visit_date':      request.form.get('visit_date', str(note['visit_date'])),
        'visit_type':      request.form.get('visit_type', note['visit_type']),
        'status':          request.form.get('status', note['status']),
        'chief_complaint': sanitise_medical_notes(request.form.get('chief_complaint','')),
        'diagnosis':       sanitise_medical_notes(request.form.get('diagnosis','')),
        'treatment':       sanitise_medical_notes(request.form.get('treatment','')),
        'clinical_notes':  sanitise_medical_notes(request.form.get('clinical_notes','')),
        'follow_up':       sanitise_medical_notes(request.form.get('follow_up','')),
    }
    db.save_visit_note(data, note_id=note_id)
    db.log_activity(session['staff_id'], 'Updated visit note',
                    f"Note #{note_id} - Patient ID {pid}")
    flash('Visit note updated.', 'success')
    return redirect(url_for('patient_view', pid=pid) + '#tab-visits')

@app.route('/patients/<int:pid>/visits/<int:note_id>/delete', methods=['POST'])
@permission_required('clinical')
def visit_note_delete(pid, note_id):
    note = db.get_visit_note(note_id)
    if note and note['patient_id'] == pid:
        db.delete_visit_note(note_id)
        db.log_activity(session['staff_id'], 'Deleted visit note',
                        f"Note #{note_id} - Patient ID {pid}")
        flash('Visit note deleted.', 'success')
    return redirect(url_for('patient_view', pid=pid) + '#tab-visits')

@app.route('/patients/<int:pid>/visits/<int:note_id>/notes', methods=['POST'])
@permission_required('clinical')
def visit_note_save_notes(pid, note_id):
    note = db.get_visit_note(note_id)
    if not note or note['patient_id'] != pid:
        return jsonify({'ok': False, 'error': 'Not found'}), 404
    clinical_notes = (request.json or {}).get('clinical_notes', '')
    data = {
        'doctor_id': note['doctor_id'], 'visit_date': str(note['visit_date']),
        'visit_type': note['visit_type'], 'status': note['status'],
        'chief_complaint': note['chief_complaint'], 'diagnosis': note['diagnosis'],
        'treatment': note['treatment'], 'clinical_notes': clinical_notes.strip() or None,
        'follow_up': note['follow_up'],
    }
    db.save_visit_note(data, note_id=note_id)
    return jsonify({'ok': True})

# ── AI: Symptom Analysis ───────────────────────────────────────────────────────

@app.route('/patients/<int:pid>/symptoms', methods=['GET', 'POST'])
@permission_required('clinical')
@sec.audit_access('symptom_analysis', id_arg='pid')
def symptom_analysis(pid):
    patient  = db.get_patient(pid)
    if not patient:
        return redirect(url_for('patients'))
    result   = None
    symptoms = ''
    if request.method == 'POST':
        symptoms = request.form.get('symptoms', '').strip()
        try:
            # ── AI Firewall: anonymise patient, validate, then call AI ─────────
            # Zero clinical data — only doctor's typed symptoms reach AI
            prompt, meta = _fw.build_safe_prompt(
                patient={},
                clinical_question=symptoms,
                prompt_type='symptom',
            )
            result = _fw.call_ai(
                prompt=prompt,
                ai_func=_gemini,
                request_ref=meta['request_ref'],
                staff_id=session['staff_id'],
                prompt_type='symptom',
            )
            db.save_symptom_analysis(pid, session['staff_id'], symptoms, result)
            db.log_activity(session['staff_id'], 'AI symptom analysis',
                            f"Patient ID {pid} (anonymised)")
        except ValueError as e:
            result = f"Bllokuar nga muri i sigurisë: {e}"
        except Exception as e:
            result = f"AI Error: {str(e)}\n\nMake sure your GEMINI_API_KEY is set in config.py"
    past = db.get_symptom_analyses(pid)
    return render_template('symptom_analysis.html', user=current_user(),
                           practice_name=PRACTICE_NAME, patient=patient,
                           result=result, symptoms=symptoms, past=past)

# ── AI: Scan Analysis ──────────────────────────────────────────────────────────

@app.route('/patients/<int:pid>/scans', methods=['GET', 'POST'])
@permission_required('clinical')
@sec.audit_access('scan_analysis', id_arg='pid')
def scan_analysis(pid):
    patient = db.get_patient(pid)
    if not patient:
        return redirect(url_for('patients'))
    sec.log_phi_access(session['staff_id'], 'scan_analysis', pid, 'VIEW')
    result = None
    if request.method == 'POST':
        scan_type        = request.form.get('scan_type', 'Medical scan')
        clinical_context = request.form.get('context', '')
        file             = request.files.get('scan_file')
        ok_upload, upload_err = validate_upload(file) if file else (False, 'No file')
        if file and ok_upload:
            filename  = secure_filename(f"p{pid}_{file.filename}")
            filepath  = os.path.join(UPLOAD_FOLDER, filename)
            file.save(filepath)
            with open(filepath, 'rb') as f:
                image_data = base64.standard_b64encode(f.read()).decode('utf-8')
            ext        = filename.rsplit('.', 1)[1].lower()
            media_map  = {'jpg':'image/jpeg','jpeg':'image/jpeg','png':'image/png',
                          'gif':'image/gif','bmp':'image/bmp'}
            media_type = media_map.get(ext, 'image/jpeg')
            # Scan: anonymise patient first, then build image prompt
            safe_p, _, _ = _fw.minimise_patient(patient)
            _, meta = _fw.build_safe_prompt(
                patient=patient,
                clinical_question=f"Scan type: {scan_type}. Context: {clinical_context or 'Not provided'}",
                prompt_type='general',
            )
            safe_prompt = (
                f"[SYSTEM INSTRUCTION]\nYou are a radiological decision-support assistant.\n"
                f"Reference ID: {meta['request_ref']}\n"
                # Zero clinical data — scan context typed by doctor only
                f"Scan type: {scan_type}. Context provided by physician.\n"
                f"Scan type: {scan_type}\nClinical context: {clinical_context or 'Not provided'}\n\n"
                f"Describe:\n1. Image modality\n2. Image quality\n3. Visible structures\n"
                f"4. Notable features vs normal anatomy\n5. Summary for reviewing physician\n"
                f"For clinical decision support only — physician makes all diagnoses."
            )
            try:
                result = _gemini(safe_prompt, image_data=image_data, media_type=media_type)
                result = _fw.call_ai(
                    prompt='__scan_passthrough__',
                    ai_func=lambda _: result,
                    request_ref=meta['request_ref'],
                    staff_id=session['staff_id'],
                    prompt_type='general',
                )
                db.save_scan_analysis(pid, session['staff_id'], scan_type, filename, result)
                db.log_activity(session['staff_id'], 'AI scan analysis',
                                f"{scan_type} — Patient ID {pid} (anonymised)")
            except ValueError as e:
                result = f"Bllokuar nga muri i sigurisë: {e}"
            except Exception as e:
                result = f"AI Error: {str(e)}\n\nMake sure your GEMINI_API_KEY is set in config.py"
        else:
            flash('Please upload a valid image file.', 'danger')
    past = db.get_scan_analyses(pid)
    return render_template('scan_analysis.html', user=current_user(),
                           practice_name=PRACTICE_NAME, patient=patient,
                           result=result, past=past)

# ── AI: Doctor's Assistant (Hero Feature) ─────────────────────────────────────

@app.route('/ai-assistant', methods=['GET', 'POST'])
@permission_required('clinical')
def ai_assistant():
    result      = None
    query       = ''
    mode        = request.form.get('mode', 'general')
    patient     = None
    scan_result = None

    if request.method == 'POST':
        query      = request.form.get('query', '').strip()
        patient_id = request.form.get('patient_id', '').strip()
        lang_instr = _lang_instruction()

        patient_context = ''
        if patient_id:
            p = db.get_patient(int(patient_id))
            if p:
                patient = p
                age = ''
                if p['dob']:
                    age = f"{(date.today() - p['dob']).days // 365} yrs"
                # Zero clinical data policy — no DB data sent to AI
                patient_context = ""  # doctor types all context manually

        # ── Scan upload mode ──────────────────────────────────────────────
        if mode == 'scan':
            scan_file    = request.files.get('scan_file')
            scan_type    = request.form.get('scan_type', 'Medical scan')
            scan_context = request.form.get('scan_context', '').strip()

            if scan_file and allowed_file(scan_file.filename):
                import base64 as _b64
                image_data = _b64.standard_b64encode(scan_file.read()).decode('utf-8')
                ext        = scan_file.filename.rsplit('.', 1)[1].lower()
                media_map  = {'jpg':'image/jpeg','jpeg':'image/jpeg','png':'image/png',
                              'gif':'image/gif','bmp':'image/bmp','tiff':'image/tiff',
                              'pdf':'application/pdf'}
                media_type = media_map.get(ext, 'image/jpeg')

                prompt = f"""Describe the visible anatomical structures and landmarks in this medical image for clinical reference.
{lang_instr}

Scan type: {scan_type}
{patient_context}
Clinical context: {scan_context or 'Not provided'}

Please describe:
1. Image modality identified
2. Image quality assessment
3. Visible structures and landmarks
4. Any notable features or asymmetries
5. Summary observations for the reviewing physician

For clinical decision support only — physician makes all diagnoses.
{_ai_suffix()}"""

                try:
                    result = _gemini(prompt, image_data=image_data, media_type=media_type)
                    if patient_id:
                        db.save_scan_analysis(int(patient_id), session['staff_id'],
                                              scan_type, scan_file.filename, result)
                    db.log_activity(session['staff_id'], 'AI Scan (Assistant)',
                                    f"{scan_type}")
                except Exception as e:
                    result = f"AI Error: {str(e)}"
            else:
                result = "Please upload a valid image file (PNG, JPG, TIFF, BMP, PDF)."
        else:
            # ── Text modes ────────────────────────────────────────────────
            mode_instructions = {
                'general':     'You are an expert AI clinical assistant helping a licensed doctor. Provide thorough, accurate, evidence-based clinical guidance.',
                'drug':        'You are an expert clinical pharmacologist. Provide detailed drug interaction analysis, contraindications, dosing adjustments, and safety warnings.',
                'differential':'You are an expert diagnostician. Provide a structured differential diagnosis with probability ranking, key distinguishing features, and recommended investigations.',
                'treatment':   'You are a clinical guidelines expert. Provide evidence-based treatment recommendations with first-line, second-line options, and monitoring parameters.',
                'summarise':   'You are a medical documentation expert. Summarise the provided clinical information clearly and concisely in structured format suitable for medical records.',
            }
            system_instruction = mode_instructions.get(mode, mode_instructions['general'])
            full_prompt = f"{system_instruction}\n{lang_instr}\n\n{patient_context}\nQuestion: {query}{_ai_suffix()}"

            try:
                result = _gemini(full_prompt)
                db.log_activity(session['staff_id'], 'AI Assistant',
                                f"Mode: {mode} | {query[:80]}")
            except Exception as e:
                result = f"AI Error: {str(e)}\n\nPlease check your API keys in config.py"

    patients_list = db.get_all_patients()
    return render_template('ai_assistant.html', user=current_user(),
                           practice_name=PRACTICE_NAME,
                           result=result, query=query, mode=mode,
                           patient=patient, patients_list=patients_list)

# ── Team ───────────────────────────────────────────────────────────────────────

@app.route('/team')
@permission_required('team_view')
def team():
    staff = db.get_all_staff()
    return render_template('team.html', user=current_user(),
                           practice_name=PRACTICE_NAME, staff=staff)

@app.route('/team/new', methods=['GET', 'POST'])
@admin_required
def staff_new():
    if request.method == 'POST':
        password = request.form.get('password', '')
        role = request.form.get('role', 'doctor')
        data = {
            'name':        request.form.get('name', '').strip(),
            'email':       request.form.get('email', '').strip(),
            'password':    bcrypt.generate_password_hash(password).decode('utf-8'),
            'role':        role,
            'specialty':   request.form.get('specialty', '').strip(),
            'phone':       request.form.get('phone', '').strip(),
            'permissions': build_permissions_json(request.form),
        }
        try:
            db.save_staff(data)
            db.log_activity(session['staff_id'], 'Added staff member', data['name'])
            flash(f"Staff member {data['name']} added.", 'success')
            return redirect(url_for('team'))
        except Exception as e:
            flash(f'Error: {e}', 'danger')
    return render_template('staff_form.html', user=current_user(),
                           practice_name=PRACTICE_NAME, member=None,
                           all_permissions=ALL_PERMISSIONS, role_presets=ROLE_PRESETS)

@app.route('/team/<int:sid>/edit', methods=['GET', 'POST'])
@admin_required
def staff_edit(sid):
    member = db.get_staff_by_id(sid)
    if not member:
        return redirect(url_for('team'))
    if request.method == 'POST':
        data = {
            'name':        request.form.get('name', '').strip(),
            'email':       request.form.get('email', '').strip(),
            'role':        request.form.get('role', 'doctor'),
            'specialty':   request.form.get('specialty', '').strip(),
            'phone':       request.form.get('phone', '').strip(),
            'active':      request.form.get('active') == 'on',
            'permissions': build_permissions_json(request.form),
        }
        db.save_staff(data, sid=sid)
        db.log_activity(session['staff_id'], 'Updated staff member', data['name'])
        flash('Staff member updated.', 'success')
        return redirect(url_for('team'))
    member_perms = parse_permissions(member.get('permissions'))
    return render_template('staff_form.html', user=current_user(),
                           practice_name=PRACTICE_NAME, member=member,
                           member_perms=member_perms,
                           all_permissions=ALL_PERMISSIONS, role_presets=ROLE_PRESETS)

@app.route('/team/<int:sid>/delete', methods=['POST'])
@admin_required
def staff_delete(sid):
    member = db.get_staff_by_id(sid)
    if member and member['id'] != session['staff_id']:
        db.delete_staff(sid)
        db.log_activity(session['staff_id'], 'Deactivated staff member', member['name'])
        flash('Staff member deactivated.', 'success')
    return redirect(url_for('team'))

# ── Calendar ───────────────────────────────────────────────────────────────────

@app.route('/calendar')
@permission_required('appointments')
def calendar():
    import calendar as cal_mod
    view     = request.args.get('view', 'week')
    date_str = request.args.get('date', date.today().isoformat())
    sel      = date.fromisoformat(date_str)

    week_days = None
    month_weeks = None

    if view == 'day':
        appointments = db.get_appointments_for_day(date_str)
        prev_date    = (sel - timedelta(days=1)).isoformat()
        next_date    = (sel + timedelta(days=1)).isoformat()

    elif view == 'month':
        # Build calendar grid for the month
        first_day = sel.replace(day=1)
        last_day  = sel.replace(day=cal_mod.monthrange(sel.year, sel.month)[1])
        # Pad to full weeks
        grid_start = first_day - timedelta(days=first_day.weekday())
        grid_end   = last_day + timedelta(days=(6 - last_day.weekday()))
        # Fetch all appointments in the grid range
        appointments = db.get_appointments_for_week(grid_start.isoformat(), grid_end.isoformat())
        prev_date    = (first_day - timedelta(days=1)).replace(day=1).isoformat()
        next_date    = (last_day + timedelta(days=1)).isoformat()
        # Build weeks list for template
        month_weeks = []
        cur = grid_start
        while cur <= grid_end:
            week = [cur + timedelta(days=i) for i in range(7)]
            month_weeks.append(week)
            cur += timedelta(days=7)

    else:  # week
        week_start   = sel - timedelta(days=sel.weekday())
        week_end     = week_start + timedelta(days=6)
        appointments = db.get_appointments_for_week(week_start.isoformat(), week_end.isoformat())
        prev_date    = (week_start - timedelta(days=7)).isoformat()
        next_date    = (week_start + timedelta(days=7)).isoformat()
        week_days    = [week_start + timedelta(days=i) for i in range(7)]

    appt_by_date = {}
    for a in appointments:
        d = a['appointment_date'].isoformat()
        appt_by_date.setdefault(d, []).append(a)

    # ── Mini calendar sidebar — always build month grid ─────────
    import calendar as _cal
    mini_first   = sel.replace(day=1)
    mini_last    = sel.replace(day=_cal.monthrange(sel.year, sel.month)[1])
    mini_start   = mini_first - timedelta(days=mini_first.weekday())
    mini_end     = mini_last  + timedelta(days=(6 - mini_last.weekday()))
    mini_cur     = mini_start
    mini_weeks   = []
    while mini_cur <= mini_end:
        mini_weeks.append([mini_cur + timedelta(days=i) for i in range(7)])
        mini_cur += timedelta(days=7)
    # Fetch any missing dates for the mini grid
    for a in db.get_appointments_for_week(mini_start.isoformat(), mini_end.isoformat()):
        d = a['appointment_date'].isoformat()
        if d not in appt_by_date:
            appt_by_date.setdefault(d, []).append(a)
    prev_month = (mini_first - timedelta(days=1)).replace(day=1).isoformat()
    next_month = (mini_last  + timedelta(days=1)).isoformat()
    if month_weeks is None:
        month_weeks = mini_weeks

    return render_template('calendar.html', user=current_user(),
                           practice_name=PRACTICE_NAME,
                           view=view, selected_date=sel, date_str=date_str,
                           appointments=appointments, appt_by_date=appt_by_date,
                           week_days=week_days, month_weeks=month_weeks,
                           prev_date=prev_date, next_date=next_date,
                           prev_month=prev_month, next_month=next_month,
                           staff=db.get_all_staff(), patients=db.get_all_patients(),
                           today=date.today().isoformat())

@app.route('/calendar/new', methods=['GET', 'POST'])
@permission_required('appointments')
def appointment_new():
    if request.method == 'POST':
        data = {
            'patient_id':       request.form.get('patient_id'),
            'doctor_id':        request.form.get('doctor_id'),
            'appointment_date': request.form.get('appointment_date'),
            'start_time':       request.form.get('start_time'),
            'duration_mins':    request.form.get('duration_mins', 30),
            'appointment_type': request.form.get('appointment_type'),
            'notes':            request.form.get('notes'),
            'status':           request.form.get('status', 'scheduled'),
        }
        db.save_appointment(data, created_by=session['staff_id'])
        db.log_activity(session['staff_id'], 'New appointment',
                        f"Patient ID {data['patient_id']} on {data['appointment_date']}")
        flash('Appointment scheduled.', 'success')
        return redirect(url_for('calendar', date=data['appointment_date']))
    # Pre-fill date/time from query params (used by calendar click-to-book)
    prefill_date = request.args.get('date', date.today().isoformat())
    prefill_time = request.args.get('time', '09:00')
    return render_template('appointment_form.html', user=current_user(),
                           practice_name=PRACTICE_NAME, appt=None,
                           staff=db.get_all_staff(), patients=db.get_all_patients(),
                           today=prefill_date,
                           prefill_time=prefill_time)

@app.route('/calendar/<int:appt_id>/edit', methods=['GET', 'POST'])
@permission_required('appointments')
def appointment_edit(appt_id):
    appt = db.get_appointment(appt_id)
    if not appt:
        return redirect(url_for('calendar'))
    if request.method == 'POST':
        data = {
            'patient_id':       request.form.get('patient_id'),
            'doctor_id':        request.form.get('doctor_id'),
            'appointment_date': request.form.get('appointment_date'),
            'start_time':       request.form.get('start_time'),
            'duration_mins':    request.form.get('duration_mins', 30),
            'appointment_type': request.form.get('appointment_type'),
            'notes':            request.form.get('notes'),
            'status':           request.form.get('status', 'scheduled'),
        }
        db.save_appointment(data, appt_id=appt_id)
        db.log_activity(session['staff_id'], 'Updated appointment', f"#{appt_id}")
        flash('Appointment updated.', 'success')
        return redirect(url_for('calendar', date=data['appointment_date']))
    return render_template('appointment_form.html', user=current_user(),
                           practice_name=PRACTICE_NAME, appt=appt,
                           staff=db.get_all_staff(), patients=db.get_all_patients(),
                           today=appt['appointment_date'].isoformat())

@app.route('/calendar/<int:appt_id>/delete', methods=['POST'])
@permission_required('appointments')
def appointment_delete(appt_id):
    appt = db.get_appointment(appt_id)
    if appt:
        db.delete_appointment(appt_id)
        db.log_activity(session['staff_id'], 'Deleted appointment', f"#{appt_id}")
        flash('Appointment deleted.', 'success')
    return redirect(url_for('calendar'))


# ── Campaigns ─────────────────────────────────────────────────────────────────

# ── Activity ───────────────────────────────────────────────────────────────────

@app.route('/activity')
@login_required
def activity():
    """Merged into /reports/audit which has filters, staff summary, and PHI highlighting."""
    return redirect(url_for('audit_report'))

# ── AI: Inline scan endpoint (used from patient Clinical tab) ─────────────────

@app.route('/api/ai-scan-inline', methods=['POST'])
@permission_required('clinical')
def api_ai_scan_inline():
    body       = request.json or {}
    patient_id = body.get('patient_id')
    scan_type  = body.get('scan_type', 'Medical scan')
    context    = body.get('context', '')
    image_data = body.get('image_data', '')
    media_type = body.get('media_type', 'image/jpeg')

    if not image_data:
        return jsonify({'ok': False, 'error': 'No image data'}), 400

    patient_ctx = ''
    if patient_id:
        p = db.get_patient(int(patient_id))
        if p:
            patient_ctx = f"Patient: {p['first_name']} {p['last_name']} | Conditions: {p.get('medical_history','None')}"

    lang_instr = _lang_instruction()
    prompt = f"""Describe the visible anatomical structures and landmarks in this medical image for clinical reference.
{lang_instr}

Scan type reported: {scan_type}
{patient_ctx}
Clinical context: {context or 'Not provided'}

Please describe:
1. Image modality identified
2. Image quality assessment
3. Visible structures and landmarks
4. Any notable features or asymmetries compared to normal
5. Summary observations for the reviewing physician

For clinical decision support only — physician makes all diagnoses.
{_ai_suffix()}"""

    try:
        result = _gemini(prompt, image_data=image_data, media_type=media_type)
        if patient_id:
            db.save_scan_analysis(int(patient_id), session['staff_id'], scan_type, 'inline', result)
        db.log_activity(session['staff_id'], 'AI inline scan', f"{scan_type}")
        return jsonify({'ok': True, 'result': result})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ── AI: Inline AJAX endpoint (for split-panel in visit notes) ─────────────────

@app.route('/api/ai-assist', methods=['POST'])
@permission_required('clinical')
def api_ai_assist():
    """
    Lightweight AJAX endpoint used by the split-panel AI assistant
    inside the visit notes tab. Returns JSON {ok, result}.
    """
    body       = request.json or {}
    query      = (body.get('query') or '').strip()
    mode       = body.get('mode', 'general')
    patient_id = body.get('patient_id')
    notes_ctx  = (body.get('notes_context') or '').strip()

    if not query:
        return jsonify({'ok': False, 'error': 'Empty query'}), 400

    patient_context = ''
    if patient_id:
        p = db.get_patient(int(patient_id))
        if p:
            age = ''
            if p['dob']:
                age = f"{(date.today() - p['dob']).days // 365} yrs"
            # Zero clinical data policy — no DB data sent to AI
            patient_context = ""  # doctor types all context manually

    notes_section = f"\nCurrent draft notes:\n{notes_ctx}\n" if notes_ctx else ''

    lang = session.get('lang', 'sq')
    lang_instruction = (
        "IMPORTANT: You MUST respond in Albanian (Shqip) language only, regardless of the language of the question."
        if lang == 'sq' else
        "Respond in English."
    )

    mode_prompts = {
        'general':      'You are an expert clinical assistant helping a licensed doctor. Be concise and practical.',
        'differential': 'You are an expert diagnostician. Give a ranked differential diagnosis with key distinguishing features.',
        'drug':         'You are a clinical pharmacologist. Give drug interactions, contraindications, and dosing guidance.',
        'treatment':    'You are a clinical guidelines expert. Give evidence-based first and second-line treatment options.',
        'summarise':    'You are a medical documentation expert. Summarise the notes below into a clean, structured clinical note.',
    }
    system = mode_prompts.get(mode, mode_prompts['general'])

    prompt = f"{system}\n{lang_instruction}\n\n{patient_context}{notes_section}\nQuestion: {query}{_ai_suffix()}"

    try:
        result = _gemini(prompt)
        db.log_activity(session['staff_id'], 'AI inline assist',
                        f"Mode:{mode} | {query[:60]}")
        return jsonify({'ok': True, 'result': result})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


# ── Inbox ──────────────────────────────────────────────────────────────────────

@app.route('/inbox')
@login_required
def inbox():
    messages = db.get_messages(session['staff_id'])
    staff    = db.get_all_staff()
    return render_template('inbox.html', user=current_user(),
                           practice_name=PRACTICE_NAME,
                           messages=messages, staff=staff)

@app.route('/inbox/send', methods=['POST'])
@login_required
def inbox_send():
    receiver_id = request.form.get('receiver_id')
    subject     = request.form.get('subject', '').strip()
    body        = request.form.get('body', '').strip()
    if receiver_id and body:
        db.send_message(session['staff_id'], int(receiver_id), subject, body)
        db.log_activity(session['staff_id'], 'Message sent', f"To staff #{receiver_id}")
        flash('Message sent.', 'success')
    return redirect(url_for('inbox'))

@app.route('/inbox/<int:msg_id>/read', methods=['POST'])
@login_required
def inbox_read(msg_id):
    db.mark_message_read(msg_id)
    return redirect(url_for('inbox'))

@app.route('/inbox/<int:msg_id>/delete', methods=['POST'])
@login_required
def inbox_delete(msg_id):
    db.delete_message(msg_id)
    flash('Message deleted.', 'success')
    return redirect(url_for('inbox'))

# ── Billing ────────────────────────────────────────────────────────────────────

@app.route('/billing')
@permission_required('billing')
def billing():
    items   = db.get_invoices()
    today   = date.today()
    return render_template('billing.html', user=current_user(),
                           practice_name=PRACTICE_NAME,
                           items=items, patient=None, today=today)

@app.route('/billing/new', methods=['GET', 'POST'])
@permission_required('billing')
def billing_new():
    if request.method == 'POST':
        data = {
            'patient_id':   request.form.get('patient_id'),
            'doctor_id':    request.form.get('doctor_id') or None,
            'invoice_date': request.form.get('invoice_date'),
            'due_date':     request.form.get('due_date') or None,
            'items':        request.form.get('items', '').strip() or None,
            'total_amount': float(request.form.get('total_amount', 0) or 0),
            'amount_paid':  float(request.form.get('amount_paid', 0) or 0),
            'status':       request.form.get('status', 'unpaid'),
            'notes':        request.form.get('notes', '').strip() or None,
        }
        inv_id = db.save_invoice(data, created_by=session['staff_id'])
        db.log_activity(session['staff_id'], 'New invoice', f"Patient #{data['patient_id']}")
        flash('Invoice created.', 'success')
        return redirect(url_for('billing'))
    return render_template('invoice_form.html', user=current_user(),
                           practice_name=PRACTICE_NAME,
                           patient=None, patients=db.get_all_patients(),
                           staff=db.get_all_staff(), today=date.today().isoformat())

@app.route('/patients/<int:pid>/billing')
@permission_required('billing')
def patient_billing(pid):
    patient = db.get_patient(pid)
    if not patient: return redirect(url_for('patients'))
    items = db.get_invoices(pid)
    today = date.today()
    return render_template('billing.html', user=current_user(),
                           practice_name=PRACTICE_NAME,
                           items=items, patient=patient, today=today)

@app.route('/patients/<int:pid>/billing/new', methods=['GET', 'POST'])
@permission_required('billing')
def patient_billing_new(pid):
    patient = db.get_patient(pid)
    if not patient: return redirect(url_for('patients'))
    if request.method == 'POST':
        data = {
            'patient_id':   pid,
            'doctor_id':    request.form.get('doctor_id') or None,
            'invoice_date': request.form.get('invoice_date'),
            'due_date':     request.form.get('due_date') or None,
            'items':        request.form.get('items', '').strip() or None,
            'total_amount': float(request.form.get('total_amount', 0) or 0),
            'amount_paid':  float(request.form.get('amount_paid', 0) or 0),
            'status':       request.form.get('status', 'unpaid'),
            'notes':        request.form.get('notes', '').strip() or None,
        }
        db.save_invoice(data, created_by=session['staff_id'])
        db.log_activity(session['staff_id'], 'New invoice', f"Patient {patient['first_name']} {patient['last_name']}")
        flash('Invoice created.', 'success')
        return redirect(url_for('patient_billing', pid=pid))
    return render_template('invoice_form.html', user=current_user(),
                           practice_name=PRACTICE_NAME,
                           patient=patient, patients=None,
                           staff=db.get_all_staff(), today=date.today().isoformat())

@app.route('/billing/<int:inv_id>/print')
@permission_required('billing')
def billing_print(inv_id):
    item = db.get_invoice(inv_id)
    if not item: return redirect(url_for('billing'))
    return render_template('invoice_print.html', item=item,
                           practice_name=PRACTICE_NAME, today=date.today())

@app.route('/billing/<int:inv_id>/pay', methods=['POST'])
@permission_required('billing')
def billing_pay(inv_id):
    db.pay_invoice(inv_id)
    db.log_activity(session['staff_id'], 'Invoice paid', f"INV-{inv_id:04d}")
    flash('Invoice marked as paid.', 'success')
    return redirect(request.referrer or url_for('billing'))

@app.route('/billing/<int:inv_id>/delete', methods=['POST'])
@permission_required('billing')
def billing_delete(inv_id):
    db.delete_invoice(inv_id)
    db.log_activity(session['staff_id'], 'Deleted invoice', f"INV-{inv_id:04d}")
    flash('Invoice deleted.', 'success')
    return redirect(request.referrer or url_for('billing'))

# ── Lab Tests ──────────────────────────────────────────────────────────────────

@app.route('/patients/<int:pid>/labs')
@permission_required('clinical')
def lab_tests(pid):
    patient = db.get_patient(pid)
    if not patient: return redirect(url_for('patients'))
    items = db.get_lab_tests(pid)
    return render_template('lab_tests.html', user=current_user(),
                           practice_name=PRACTICE_NAME, patient=patient, items=items)

@app.route('/patients/<int:pid>/labs/new', methods=['GET', 'POST'])
@permission_required('clinical')
def lab_test_new(pid):
    patient = db.get_patient(pid)
    if not patient: return redirect(url_for('patients'))
    if request.method == 'POST':
        data = {
            'patient_id':   pid,
            'doctor_id':    session['staff_id'],
            'test_name':    request.form.get('test_name', '').strip(),
            'test_type':    request.form.get('test_type'),
            'ordered_date': request.form.get('ordered_date'),
            'status':       request.form.get('status', 'ordered'),
            'notes':        request.form.get('notes', '').strip() or None,
        }
        db.save_lab_test(data, created_by=session['staff_id'])
        db.log_activity(session['staff_id'], 'Lab test ordered', f"{data['test_name']} — {patient['first_name']} {patient['last_name']}")
        flash('Lab test ordered.', 'success')
        return redirect(url_for('lab_tests', pid=pid))
    return render_template('lab_test_form.html', user=current_user(),
                           practice_name=PRACTICE_NAME, patient=patient,
                           today=date.today().isoformat())

@app.route('/labs/<int:lab_id>/result', methods=['POST'])
@permission_required('clinical')
def lab_result(lab_id):
    db.update_lab_result(
        lab_id,
        request.form.get('result_date'),
        request.form.get('result_value', '').strip(),
        request.form.get('reference_range', '').strip() or None,
        'is_abnormal' in request.form
    )
    db.log_activity(session['staff_id'], 'Lab result entered', f"Lab #{lab_id}")
    flash('Result saved.', 'success')
    return redirect(request.referrer or url_for('patients'))

@app.route('/labs/<int:lab_id>/delete', methods=['POST'])
@permission_required('clinical')
def lab_test_delete(lab_id):
    db.delete_lab_test(lab_id)
    flash('Lab test deleted.', 'success')
    return redirect(request.referrer or url_for('patients'))

# ── Patient Documents ──────────────────────────────────────────────────────────

@app.route('/patients/<int:pid>/documents')
@permission_required('patients_view')
def patient_documents(pid):
    patient = db.get_patient(pid)
    if not patient: return redirect(url_for('patients'))
    docs = db.get_patient_documents(pid)
    return render_template('patient_documents.html', user=current_user(),
                           practice_name=PRACTICE_NAME, patient=patient, docs=docs)

@app.route('/patients/<int:pid>/documents/upload', methods=['POST'])
@permission_required('patients_edit')
def patient_documents_upload(pid):
    patient = db.get_patient(pid)
    if not patient: return redirect(url_for('patients'))
    file = request.files.get('doc_file')
    if file and allowed_file(file.filename):
        filename = secure_filename(f"doc_p{pid}_{file.filename}")
        file.save(os.path.join(UPLOAD_FOLDER, filename))
        data = {
            'patient_id': pid,
            'doc_type':   request.form.get('doc_type'),
            'title':      request.form.get('title', '').strip(),
            'filename':   filename,
            'notes':      request.form.get('notes', '').strip() or None,
        }
        db.save_patient_document(data, created_by=session['staff_id'])
        db.log_activity(session['staff_id'], 'Document uploaded', f"{data['title']} — {patient['first_name']} {patient['last_name']}")
        flash('Document uploaded.', 'success')
    else:
        flash('Invalid file type.', 'danger')
    return redirect(url_for('patient_documents', pid=pid))

@app.route('/documents/<int:doc_id>/delete', methods=['POST'])
@permission_required('patients_edit')
def patient_document_delete(doc_id):
    db.delete_patient_document(doc_id)
    flash('Document deleted.', 'success')
    return redirect(request.referrer or url_for('patients'))

# ── Account History ────────────────────────────────────────────────────────────

@app.route('/patients/<int:pid>/account-history')
@permission_required('billing')
def patient_account_history(pid):
    patient  = db.get_patient(pid)
    if not patient: return redirect(url_for('patients'))
    invoices = db.get_invoices(pid)
    return render_template('account_history.html', user=current_user(),
                           practice_name=PRACTICE_NAME, patient=patient,
                           invoices=invoices)

# ── Address Book ───────────────────────────────────────────────────────────────

@app.route('/address-book')
@login_required
def address_book():
    entries = db.get_address_book()
    return render_template('address_book.html', user=current_user(),
                           practice_name=PRACTICE_NAME, entries=entries)

@app.route('/address-book/new', methods=['GET', 'POST'])
@login_required
def address_book_new():
    if request.method == 'POST':
        data = {k: v.strip() or None for k, v in request.form.items()}
        db.save_address_book_entry(data)
        db.log_activity(session['staff_id'], 'New address book entry', data.get('name',''))
        flash('Contact added.', 'success')
        return redirect(url_for('address_book'))
    return render_template('address_book_form.html', user=current_user(),
                           practice_name=PRACTICE_NAME, entry=None)

@app.route('/address-book/<int:eid>/edit', methods=['GET', 'POST'])
@login_required
def address_book_edit(eid):
    entry = db.get_address_book_entry(eid)
    if not entry: return redirect(url_for('address_book'))
    if request.method == 'POST':
        data = {k: v.strip() or None for k, v in request.form.items()}
        db.save_address_book_entry(data, entry_id=eid)
        flash('Contact updated.', 'success')
        return redirect(url_for('address_book'))
    return render_template('address_book_form.html', user=current_user(),
                           practice_name=PRACTICE_NAME, entry=entry)

@app.route('/address-book/<int:eid>/delete', methods=['POST'])
@login_required
def address_book_delete(eid):
    db.delete_address_book_entry(eid)
    flash('Contact deleted.', 'success')
    return redirect(url_for('address_book'))

# ── Appointment Reminders & Slip ───────────────────────────────────────────────

# ── Consultation Notes ─────────────────────────────────────────────────────────

@app.route('/patients/<int:pid>/notes')
@permission_required('clinical')
def consultation_notes(pid):
    patient = db.get_patient(pid)
    if not patient: return redirect(url_for('patients'))
    notes = db.get_consultation_notes(pid)
    return render_template('consultation_notes.html', user=current_user(),
                           practice_name=PRACTICE_NAME, patient=patient, notes=notes)

@app.route('/patients/<int:pid>/notes/new', methods=['GET', 'POST'])
@permission_required('clinical')
def consultation_note_new(pid):
    patient = db.get_patient(pid)
    if not patient: return redirect(url_for('patients'))
    if request.method == 'POST':
        data = {
            'patient_id': pid,
            'doctor_id':  request.form.get('doctor_id') or session['staff_id'],
            'visit_date': request.form.get('visit_date'),
            'subjective': request.form.get('subjective', '').strip() or None,
            'objective':  request.form.get('objective', '').strip() or None,
            'assessment': request.form.get('assessment', '').strip() or None,
            'plan':       request.form.get('plan', '').strip() or None,
            'notes':      request.form.get('notes', '').strip() or None,
        }
        db.save_consultation_note(data)
        db.log_activity(session['staff_id'], 'New consultation note', f"{patient['first_name']} {patient['last_name']}")
        flash('Consultation note saved.', 'success')
        return redirect(url_for('consultation_notes', pid=pid))
    return render_template('consultation_note_form.html', user=current_user(),
                           practice_name=PRACTICE_NAME, patient=patient, note=None,
                           staff=db.get_all_staff(), today=date.today().isoformat())

@app.route('/notes/<int:note_id>/edit', methods=['GET', 'POST'])
@permission_required('clinical')
def consultation_note_edit(note_id):
    note = db.get_consultation_note(note_id)
    if not note: return redirect(url_for('patients'))
    patient = db.get_patient(note['patient_id'])
    if request.method == 'POST':
        data = {
            'doctor_id':  request.form.get('doctor_id') or note['doctor_id'],
            'visit_date': request.form.get('visit_date', str(note['visit_date'])),
            'subjective': request.form.get('subjective', '').strip() or None,
            'objective':  request.form.get('objective', '').strip() or None,
            'assessment': request.form.get('assessment', '').strip() or None,
            'plan':       request.form.get('plan', '').strip() or None,
            'notes':      request.form.get('notes', '').strip() or None,
        }
        db.save_consultation_note(data, note_id=note_id)
        flash('Note updated.', 'success')
        return redirect(url_for('consultation_notes', pid=patient['id']))
    return render_template('consultation_note_form.html', user=current_user(),
                           practice_name=PRACTICE_NAME, patient=patient, note=note,
                           staff=db.get_all_staff(), today=date.today().isoformat())

@app.route('/notes/<int:note_id>/delete', methods=['POST'])
@permission_required('clinical')
def consultation_note_delete(note_id):
    note = db.get_consultation_note(note_id)
    pid = note['patient_id'] if note else None
    db.delete_consultation_note(note_id)
    flash('Note deleted.', 'success')
    return redirect(url_for('consultation_notes', pid=pid) if pid else url_for('patients'))


# ── Prescriptions ─────────────────────────────────────────────────────────────

@app.route('/patients/<int:pid>/prescriptions')
@permission_required('clinical')
def prescriptions(pid):
    patient = db.get_patient(pid)
    if not patient: return redirect(url_for('patients'))
    items = db.get_prescriptions(pid)
    return render_template('prescriptions.html', user=current_user(),
                           practice_name=PRACTICE_NAME, patient=patient, items=items)

@app.route('/patients/<int:pid>/prescriptions/new', methods=['GET', 'POST'])
@permission_required('clinical')
def prescription_new(pid):
    patient = db.get_patient(pid)
    if not patient: return redirect(url_for('patients'))
    if request.method == 'POST':
        data = {
            'patient_id':   pid,
            'doctor_id':    session['staff_id'],
            'drug_name':    request.form.get('drug_name', '').strip(),
            'dosage':       request.form.get('dosage', '').strip() or None,
            'frequency':    request.form.get('frequency'),
            'route':        request.form.get('route', 'Oral'),
            'start_date':   request.form.get('start_date') or None,
            'end_date':     request.form.get('end_date') or None,
            'repeats':      request.form.get('repeats', 0),
            'instructions': request.form.get('instructions', '').strip() or None,
            'status':       request.form.get('status', 'active'),
        }
        db.save_prescription(data, created_by=session['staff_id'])
        db.log_activity(session['staff_id'], 'New prescription',
                        f"{data['drug_name']} — {patient['first_name']} {patient['last_name']}")
        flash('Prescription issued.', 'success')
        return redirect(url_for('prescriptions', pid=pid))
    return render_template('prescription_form.html', user=current_user(),
                           practice_name=PRACTICE_NAME, patient=patient,
                           today=date.today().isoformat())

@app.route('/prescriptions/<int:rx_id>/print')
@permission_required('clinical')
def prescription_print(rx_id):
    rx = db.get_prescription(rx_id)
    if not rx: return redirect(url_for('patients'))
    return render_template('prescription_print.html', rx=rx,
                           practice_name=PRACTICE_NAME, today=date.today())

@app.route('/prescriptions/<int:rx_id>/delete', methods=['POST'])
@permission_required('clinical')
def prescription_delete(rx_id):
    rx = db.get_prescription(rx_id)
    pid = rx['patient_id'] if rx else None
    db.delete_prescription(rx_id)
    flash('Prescription deleted.', 'success')
    return redirect(url_for('prescriptions', pid=pid) if pid else url_for('patients'))

# ── Recalls ────────────────────────────────────────────────────────────────────

# ── Referrals ──────────────────────────────────────────────────────────────────

@app.route('/patients/<int:pid>/referrals')
@permission_required('clinical')
def referrals(pid):
    patient = db.get_patient(pid)
    if not patient: return redirect(url_for('patients'))
    items = db.get_referrals(pid)
    return render_template('referrals.html', user=current_user(),
                           practice_name=PRACTICE_NAME, patient=patient, items=items)

@app.route('/patients/<int:pid>/referrals/new', methods=['GET', 'POST'])
@permission_required('clinical')
def referral_new(pid):
    patient = db.get_patient(pid)
    if not patient: return redirect(url_for('patients'))
    doctor = db.get_staff_by_id(session['staff_id'])
    if request.method == 'POST':
        data = {
            'patient_id':     pid,
            'doctor_id':      session['staff_id'],
            'referred_to':    request.form.get('referred_to', '').strip(),
            'specialty':      request.form.get('specialty', '').strip() or None,
            'reason':         request.form.get('reason', '').strip(),
            'urgency':        request.form.get('urgency', 'routine'),
            'letter_content': request.form.get('letter_content', '').strip() or None,
            'status':         request.form.get('status', 'draft'),
        }
        ref_id = db.save_referral(data)
        db.log_activity(session['staff_id'], 'New referral',
                        f"To {data['referred_to']} — {patient['first_name']} {patient['last_name']}")
        flash('Referral letter saved.', 'success')
        return redirect(url_for('referral_print', ref_id=ref_id))
    return render_template('referral_form.html', user=current_user(),
                           practice_name=PRACTICE_NAME, patient=patient, doctor=doctor)

@app.route('/referrals/<int:ref_id>/print')
@permission_required('clinical')
def referral_print(ref_id):
    ref = db.get_referral(ref_id)
    if not ref: return redirect(url_for('patients'))
    return render_template('referral_print.html', ref=ref,
                           practice_name=PRACTICE_NAME, today=date.today())

@app.route('/referrals/<int:ref_id>/delete', methods=['POST'])
@permission_required('clinical')
def referral_delete(ref_id):
    ref = db.get_referral(ref_id)
    pid = ref['patient_id'] if ref else None
    db.delete_referral(ref_id)
    flash('Referral deleted.', 'success')
    return redirect(url_for('referrals', pid=pid) if pid else url_for('patients'))

# ── Reminders ──────────────────────────────────────────────────────────────────

# ── Patient archive & merge ────────────────────────────────────────────────────

@app.route('/patients/archived')
@permission_required('patients_view')
def patients_archived():
    patients = db.get_archived_patients()
    return render_template('patients_archived.html', user=current_user(),
                           practice_name=PRACTICE_NAME, patients=patients)

@app.route('/patients/<int:pid>/archive', methods=['POST'])
@permission_required('patients_edit')
def patient_archive(pid):
    p = db.get_patient(pid)
    if p:
        db.archive_patient(pid)
        db.log_activity(session['staff_id'], 'Archived patient',
                        f"{p['first_name']} {p['last_name']}")
        flash('Patient archived.', 'success')
    return redirect(url_for('patients'))

@app.route('/patients/<int:pid>/unarchive', methods=['POST'])
@permission_required('patients_edit')
def patient_unarchive(pid):
    db.unarchive_patient(pid)
    db.log_activity(session['staff_id'], 'Unarchived patient', f"ID {pid}")
    flash('Patient restored to active.', 'success')
    return redirect(url_for('patients_archived'))

@app.route('/patients/duplicates')
@permission_required('patients_view')
def patients_duplicates():
    dupes = db.get_duplicate_patients()
    return render_template('patients_duplicates.html', user=current_user(),
                           practice_name=PRACTICE_NAME, dupes=dupes)

@app.route('/patients/merge', methods=['POST'])
@permission_required('patients_edit')
def patients_merge():
    keep_id  = int(request.form.get('keep_id'))
    merge_id = int(request.form.get('merge_id'))
    db.merge_patients(keep_id, merge_id)
    db.log_activity(session['staff_id'], 'Merged patients',
                    f"Kept #{keep_id}, archived #{merge_id}")
    flash('Patients merged successfully.', 'success')
    return redirect(url_for('patients_duplicates'))

# ── Reports ────────────────────────────────────────────────────────────────────

@app.route('/reports')
@permission_required('billing')
def reports():
    staff_list = db.get_all_staff()
    return render_template('reports.html', user=current_user(),
                           practice_name=PRACTICE_NAME,
                           staff=staff_list, today=date.today().isoformat())

@app.route('/reports/earnings-charges')
@permission_required('billing')
def report_earnings_charges():
    start     = request.args.get('start', date.today().replace(day=1).isoformat())
    end       = request.args.get('end',   date.today().isoformat())
    doctor_id = request.args.get('doctor_id') or None
    rows = db.report_earnings_charges(start, end, doctor_id)
    total       = sum(r['total_amount'] or 0 for r in rows)
    paid        = sum(r['amount_paid']  or 0 for r in rows)
    outstanding = sum(r['outstanding']  or 0 for r in rows)
    return render_template('report_earnings_charges.html', user=current_user(),
                           practice_name=PRACTICE_NAME,
                           rows=rows, start=start, end=end,
                           doctor_id=doctor_id, staff=db.get_all_staff(),
                           total=total, paid=paid, outstanding=outstanding)

@app.route('/reports/earnings-payments')
@permission_required('billing')
def report_earnings_payments():
    start     = request.args.get('start', date.today().replace(day=1).isoformat())
    end       = request.args.get('end',   date.today().isoformat())
    doctor_id = request.args.get('doctor_id') or None
    rows  = db.report_earnings_payments(start, end, doctor_id)
    total = sum(r['amount_paid'] or 0 for r in rows)
    return render_template('report_earnings_payments.html', user=current_user(),
                           practice_name=PRACTICE_NAME,
                           rows=rows, start=start, end=end,
                           doctor_id=doctor_id, staff=db.get_all_staff(), total=total)

@app.route('/reports/overdue')
@permission_required('billing')
def report_overdue():
    rows  = db.report_overdue()
    total = sum(r['outstanding'] or 0 for r in rows)
    return render_template('report_overdue.html', user=current_user(),
                           practice_name=PRACTICE_NAME, rows=rows, total=total)

@app.route('/reports/booking')
@permission_required('appointments')
def report_booking():
    start     = request.args.get('start', date.today().isoformat())
    end       = request.args.get('end',   date.today().isoformat())
    doctor_id = request.args.get('doctor_id') or None
    rows = db.report_booking(start, end, doctor_id)
    return render_template('report_booking.html', user=current_user(),
                           practice_name=PRACTICE_NAME,
                           rows=rows, start=start, end=end,
                           doctor_id=doctor_id, staff=db.get_all_staff())

@app.route('/reports/daily-banking')
@permission_required('billing')
def report_daily_banking():
    report_date = request.args.get('date', date.today().isoformat())
    rows  = db.report_daily_banking(report_date)
    total = sum(r['amount_paid'] or 0 for r in rows)
    return render_template('report_daily_banking.html', user=current_user(),
                           practice_name=PRACTICE_NAME,
                           rows=rows, report_date=report_date, total=total)

@app.route('/reports/transactions')
@permission_required('billing')
def report_transactions():
    start     = request.args.get('start', date.today().replace(day=1).isoformat())
    end       = request.args.get('end',   date.today().isoformat())
    doctor_id = request.args.get('doctor_id') or None
    rows  = db.report_transactions(start, end, doctor_id)
    total = sum(r['total_amount'] or 0 for r in rows)
    return render_template('report_transactions.html', user=current_user(),
                           practice_name=PRACTICE_NAME,
                           rows=rows, start=start, end=end,
                           doctor_id=doctor_id, staff=db.get_all_staff(), total=total)


# ── Vaccinations ──────────────────────────────────────────────────────────────

@app.route('/patients/<int:pid>/vaccinations')
@permission_required('clinical')
def vaccinations(pid):
    patient = db.get_patient(pid)
    if not patient: return redirect(url_for('patients'))
    items = db.get_vaccinations(pid)
    return render_template('vaccinations.html', user=current_user(),
                           practice_name=PRACTICE_NAME, patient=patient,
                           items=items, today=date.today())

@app.route('/patients/<int:pid>/vaccinations/new', methods=['POST'])
@permission_required('clinical')
def vaccination_new(pid):
    patient = db.get_patient(pid)
    if not patient: return redirect(url_for('patients'))
    data = {
        'patient_id':    pid,
        'vaccine_name':  request.form.get('vaccine_name', '').strip(),
        'dose':          request.form.get('dose'),
        'date_given':    request.form.get('date_given'),
        'batch_number':  request.form.get('batch_number', '').strip(),
        'next_due_date': request.form.get('next_due_date') or None,
        'notes':         request.form.get('notes', '').strip() or None,
    }
    db.save_vaccination(data, created_by=session['staff_id'])
    db.log_activity(session['staff_id'], 'Vaccination recorded',
                    f"{data['vaccine_name']} — {patient['first_name']} {patient['last_name']}")
    flash('Vaccination recorded.', 'success')
    return redirect(url_for('vaccinations', pid=pid))

@app.route('/vaccinations/<int:vac_id>/delete', methods=['POST'])
@permission_required('clinical')
def vaccination_delete(vac_id):
    db.delete_vaccination(vac_id)
    flash('Vaccination deleted.', 'success')
    return redirect(request.referrer or url_for('patients'))

# ── Waiting Room ───────────────────────────────────────────────────────────────

@app.route('/waiting-room')
@login_required
def waiting_room():
    queue    = db.get_room_status()   # uses the richer query with patient/doctor names
    patients = db.get_all_patients()
    staff    = db.get_all_staff()
    return render_template('waiting_room.html', user=current_user(),
                           practice_name=PRACTICE_NAME,
                           queue=queue, patients=patients, staff=staff,
                           now=datetime.now())

@app.route('/waiting-room/checkin', methods=['POST'])
@login_required
def waiting_room_checkin():
    patient_id = request.form.get('patient_id')
    doctor_id  = request.form.get('doctor_id') or None
    reason     = request.form.get('reason', '').strip() or None
    if patient_id:
        db.checkin_patient(patient_id, doctor_id, reason)
        db.log_activity(session['staff_id'], 'Patient checked in',
                        f"Patient #{patient_id}")
        flash('Patient checked in.', 'success')
    return redirect(url_for('waiting_room'))

@app.route('/waiting-room/<int:wid>/status', methods=['POST'])
@login_required
def waiting_room_status(wid):
    status = request.form.get('status', 'waiting')
    db.update_waiting_status(wid, status)
    return redirect(url_for('waiting_room'))

@app.route('/waiting-room/<int:wid>/remove', methods=['POST'])
@login_required
def waiting_room_remove(wid):
    db.remove_from_waiting(wid)
    flash('Patient discharged from waiting room.', 'success')
    return redirect(url_for('waiting_room'))


# ══════════════════════════════════════════════════════════════════════════════
#  FEATURES: Tasks, Kiosk, Room Status, Audit Report, Med Checker, Education
# ════════════════════════════════════════════
#  FEATURES: Tasks, Kiosk, Room, Audit, AI
# ════════════════════════════════════════════

@app.route('/tasks')
@login_required
def tasks():
    my_tasks  = db.get_tasks(assigned_to=session['staff_id'])
    all_tasks = db.get_tasks() if session.get('role') == 'admin' else []
    return render_template('tasks.html', user=current_user(),
        practice_name=PRACTICE_NAME, my_tasks=my_tasks, all_tasks=all_tasks,
        staff=db.get_all_staff(), patients=db.get_all_patients(), today=date.today())

@app.route('/tasks/new', methods=['POST'])
@login_required
def task_new():
    data = {'patient_id': request.form.get('patient_id') or None,
            'created_by': session['staff_id'],
            'assigned_to': request.form.get('assigned_to') or None,
            'title': request.form.get('title','').strip(),
            'description': request.form.get('description','').strip() or None,
            'priority': request.form.get('priority','normal'),
            'task_type': request.form.get('task_type','general'),
            'due_date': request.form.get('due_date') or None,
            'status': 'pending'}
    if data.get('title'):
        db.save_task(data)
        flash('Task created.', 'success')
    return redirect(request.referrer or '/tasks')

@app.route('/tasks/<int:tid>/status', methods=['POST'])
@login_required
def task_status(tid):
    task = db.get_task(tid)
    if task:
        db.save_task({'title': task['title'],
                      'status': request.form.get('status','completed'),
                      'priority': task['priority'],
                      'task_type': task['task_type']}, task_id=tid)
    return redirect(request.referrer or '/tasks')

@app.route('/tasks/<int:tid>/delete', methods=['POST'])
@login_required
def task_delete(tid):
    task = db.get_task(tid)
    if task:
        db.delete_task(tid)
        flash('Task deleted.', 'success')
    return redirect(request.referrer or '/tasks')

@app.route('/kiosk')
def kiosk():
    return render_template('kiosk.html', practice_name=PRACTICE_NAME,
        patients=db.get_all_patients(), staff=db.get_all_staff(), today=date.today())

@app.route('/kiosk/checkin', methods=['POST'])
def kiosk_checkin():
    pid    = request.form.get('patient_id')
    doc_id = request.form.get('doctor_id') or None
    reason = request.form.get('reason','').strip() or None
    note   = request.form.get('notes','').strip() or None
    if pid:
        db.checkin_patient(pid, doc_id, reason)
        if reason:
            cn = ('Patient self-reported: ' + note) if note else None
            db.save_visit_note({'patient_id': pid, 'doctor_id': doc_id,
                'visit_date': date.today().isoformat(), 'visit_type': 'Consultation',
                'status': 'scheduled', 'chief_complaint': reason,
                'clinical_notes': cn}, created_by=None)
        db.log_activity(None, 'Kiosk check-in', 'Patient #' + str(pid))
    return render_template('kiosk_confirmed.html', practice_name=PRACTICE_NAME)

@app.route('/room-status')
@login_required
def room_status():
    """Merged into /waiting-room."""
    return redirect(url_for('waiting_room'))

@app.route('/reports/audit')
@permission_required('activity_log')
def audit_report():
    days     = int(request.args.get('days', 30))
    staff_id = int(request.args.get('staff_id')) if request.args.get('staff_id') else None
    return render_template('audit_report.html', user=current_user(),
        practice_name=PRACTICE_NAME, logs=db.get_audit_report(days, staff_id),
        summary=db.get_audit_summary(days), staff=db.get_all_staff(),
        days=days, filter_staff=staff_id)

@app.route('/api/med-check', methods=['POST'])
@permission_required('clinical')
def api_med_check():
    body     = request.json or {}
    pid      = body.get('patient_id')
    new_drug = body.get('new_drug','').strip()
    if not pid or not new_drug:
        return jsonify({'ok': False, 'error': 'Missing fields'}), 400
    patient = db.get_patient(int(pid))
    if not patient:
        return jsonify({'ok': False, 'error': 'Patient not found'}), 404
    lang_instr = _lang_instruction()
    # Zero clinical data policy — only the drug name typed by doctor is sent
    prompt = ('Clinical pharmacist. ' + lang_instr + ' Respond ONLY with valid JSON, no markdown.\n'
              'New medication: ' + new_drug + '\n'
              'Provide general pharmacology information for this drug only.\n'
              'JSON: {"safe":true,"allergy_alert":null,"interactions":[],"contraindications":[],"recommendation":""}')
    try:
        msg_text__ = _gemini(prompt)
        text = msg_text__
        if text.startswith('```'): text = text.split('```')[1].lstrip('json').strip()
        result = json.loads(text)
        db.log_activity(session['staff_id'], 'Med interaction check', new_drug)
        return jsonify({'ok': True, 'result': result})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/education-resources', methods=['POST'])
@permission_required('clinical')
def api_education_resources():
    body      = request.json or {}
    diagnosis = body.get('diagnosis','').strip()
    if not diagnosis:
        return jsonify({'ok': False, 'error': 'No diagnosis'}), 400

    prompt = (
        'You are a medical educator. Generate patient education material in Albanian language.\n'
        'Topic: ' + diagnosis + '\n\n'
        'Respond ONLY with valid JSON, no markdown, no extra text.\n'
        'JSON structure:\n'
        '{'
        '"title": "short title in Albanian",'
        '"lifestyle_intro": "2-3 warm encouraging sentences in Albanian about living well with these conditions",'
        '"key_points": ["practical tip 1 in Albanian", "tip 2", "tip 3", "tip 4", "tip 5"],'
        '"when_to_seek_help": ["warning sign 1 in Albanian", "sign 2", "sign 3"],'
        '"doctor_notes": ""'
        '}'
    )

    try:
        text = _gemini(prompt)
        if '```' in text:
            text = text.split('```')[1]
            if text.startswith('json'): text = text[4:]
        text = text.strip()
        result = json.loads(text)
        return jsonify({'ok': True, 'result': result})
    except Exception as e:
        # Fallback to Groq if Gemini fails
        try:
            import os as _os3
            client = _GroqClient(api_key=_os3.environ.get('GROQ_API_KEY', GROQ_API_KEY))
            response = client.chat.completions.create(
                model='llama-3.3-70b-versatile', max_tokens=1200,
                messages=[
                    {'role': 'system', 'content': 'You are a medical educator. Respond ONLY with valid JSON, no markdown.'},
                    {'role': 'user', 'content': prompt}
                ]
            )
            text = response.choices[0].message.content.strip()
            if '```' in text:
                text = text.split('```')[1]
                if text.startswith('json'): text = text[4:]
            result = json.loads(text.strip())
            return jsonify({'ok': True, 'result': result})
        except Exception as e2:
            return jsonify({'ok': False, 'error': str(e2)}), 500


# ════════════════════════════════════════════
#  ADVANCED FEATURES: BTG, Signing, Risk, FHIR
# ════════════════════════════════════════════

# ── IP Geofencing middleware ──────────────────────────────────────────────────
def check_geofence():
    from config import ALLOWED_IPS
    if not ALLOWED_IPS: return True  # empty list = allow all
    ip = request.remote_addr or ''
    # Allow localhost always
    if ip in ('127.0.0.1', '::1', 'localhost'): return True
    return any(ip.startswith(prefix) for prefix in ALLOWED_IPS)

# ── Break-the-Glass ───────────────────────────────────────────────────────────

@app.route('/patients/<int:pid>/btg', methods=['POST'])
@login_required
def btg_request_access(pid):
    justification = request.form.get('justification','').strip()
    if not justification or len(justification) < 20:
        flash('Justification must be at least 20 characters.', 'danger')
        return redirect(url_for('patient_view', pid=pid))
    btg_id = db.btg_request(session['staff_id'], pid, justification)
    patient = db.get_patient(pid)
    pname   = patient['first_name'] + ' ' + patient['last_name'] if patient else str(pid)
    db.log_activity(session['staff_id'], 'BTG — Emergency Access Granted',
                    'Patient: ' + pname + ' | Justification: ' + justification[:100])
    # Queue admin alert (picked up by before_request)
    db.btg_mark_notified(btg_id)
    # Send internal message to all admins
    admins = [s for s in db.get_all_staff() if s['role'] == 'admin' and s['active']]
    for admin in admins:
        db.send_message(session['staff_id'], admin['id'],
                        'URGENT: Break-the-Glass Access — ' + pname,
                        session['staff_name'] + ' has invoked emergency access to patient ' +
                        pname + '.\n\nJustification: ' + justification +
                        '\n\nThis access expires in 60 minutes. Please review the audit log.')
    flash('Emergency access granted for 60 minutes. This has been logged and the administrator has been notified.', 'warning')
    return redirect(url_for('patient_view', pid=pid))

@app.route('/btg-log')
@permission_required('activity_log')
def btg_log():
    logs = db.btg_get_log(100)
    return render_template('btg_log.html', user=current_user(),
                           practice_name=PRACTICE_NAME, logs=logs)

# ── Cryptographic Document Signing ────────────────────────────────────────────

@app.route('/patients/<int:pid>/visits/<int:note_id>/sign', methods=['POST'])
@permission_required('clinical')
def sign_visit_note(pid, note_id):
    note = db.get_visit_note(note_id)
    if not note or note['patient_id'] != pid:
        return jsonify({'ok': False, 'error': 'Not found'}), 404
    content = '|'.join([
        str(note_id), str(pid),
        str(note.get('visit_date','')),
        str(note.get('diagnosis','')),
        str(note.get('clinical_notes','')),
    ])
    content_hash, sig_id = db.sign_document('visit_note', note_id, content, session['staff_id'])
    db.log_activity(session['staff_id'], 'PHI SIGN',
                    'Visit note #' + str(note_id) + ' signed — hash ' + content_hash[:16] + '...')
    return jsonify({'ok': True, 'hash': content_hash, 'sig_id': sig_id})

@app.route('/patients/<int:pid>/visits/<int:note_id>/verify')
@permission_required('clinical')
def verify_visit_note(pid, note_id):
    note = db.get_visit_note(note_id)
    if not note:
        return jsonify({'ok': False, 'error': 'Not found'}), 404
    content = '|'.join([
        str(note_id), str(pid),
        str(note.get('visit_date','')),
        str(note.get('diagnosis','')),
        str(note.get('clinical_notes','')),
    ])
    sig, intact = db.verify_document('visit_note', note_id, content)
    if not sig:
        return jsonify({'ok': True, 'signed': False, 'intact': None})
    return jsonify({'ok': True, 'signed': True, 'intact': intact,
                    'signed_by': sig.get('signed_by_name'),
                    'signed_at': str(sig.get('signed_at')),
                    'hash': sig['content_hash'][:16] + '...'})

# ── AI Risk Score ──────────────────────────────────────────────────────────────


# ── Gap-in-Care Alerts ─────────────────────────────────────────────────────────

@app.route('/api/gap-in-care/<int:pid>')
@permission_required('clinical')
def api_gap_in_care(pid):
    try:
        data = db.get_gap_in_care_data(pid)
        if not data:
            return jsonify({'ok': False, 'error': 'Not found'}), 404

        # Safely coerce all date values — psycopg2 may return date or datetime
        def to_date(val):
            if val is None:
                return None
            if hasattr(val, 'date'):   # datetime → date
                return val.date()
            return val                 # already a date

        dob       = to_date(data.get('dob'))
        last_flu  = to_date(data.get('last_flu_shot'))
        today     = date.today()

        age = int((today - dob).days // 365) if dob else 0
        gender  = str(data.get('gender') or '').lower()
        history = str(data.get('medical_history') or '').lower()
        try:
            abnormal_labs = int(data.get('abnormal_labs') or 0)
        except (TypeError, ValueError):
            abnormal_labs = 0

        alerts = []

        if not last_flu or (today - last_flu).days > 365:
            alerts.append({'type': 'vaccine', 'priority': 'routine',
                           'message': 'Vaksinimi vjetor kundër gripit i nevojshëm', 'icon': 'shield-plus'})
        if age >= 50 and 'colonoscopy' not in history:
            alerts.append({'type': 'screening', 'priority': 'high',
                           'message': 'Screening për kancerin e zorrës së trashë (moshe 50+)', 'icon': 'clipboard2-pulse'})
        if age >= 40 and 'female' in gender and 'mammogram' not in history:
            alerts.append({'type': 'screening', 'priority': 'high',
                           'message': 'Mamografia e rekomanduar (femër 40+)', 'icon': 'heart-pulse'})
        if age >= 65:
            alerts.append({'type': 'vaccine', 'priority': 'routine',
                           'message': 'Vaksinimi pneumokokal i nevojshëm (moshe 65+)', 'icon': 'shield-check'})
        if 'diabetes' in history or ' dm ' in history or history.startswith('dm') or 'diabet' in history:
            alerts.append({'type': 'monitoring', 'priority': 'high',
                           'message': 'Kontroll HbA1c — menaxhim i diabetit', 'icon': 'droplet'})
        if 'hypertension' in history or 'htn' in history or 'hipertension' in history or 'presion' in history:
            alerts.append({'type': 'monitoring', 'priority': 'moderate',
                           'message': 'Kontroll i presionit të gjakut i rekomanduar', 'icon': 'activity'})
        if age >= 45 and 'male' in gender and 'female' not in gender and 'psa' not in history and 'mashkull' in gender.lower() or (age >= 45 and gender.lower() in ['male', 'mashkull'] and 'psa' not in history):
            alerts.append({'type': 'screening', 'priority': 'routine',
                           'message': 'Diskutim PSA (mashkull 45+)', 'icon': 'person-check'})
        if abnormal_labs > 0:
            alerts.append({'type': 'urgent', 'priority': 'urgent',
                           'message': str(abnormal_labs) + ' rezultat laboratori jonormal kërkon rishikim',
                           'icon': 'exclamation-triangle'})

        return jsonify({'ok': True, 'alerts': alerts, 'count': len(alerts)})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

# ── FHIR R4 Export ────────────────────────────────────────────────────────────

@app.route('/patients/<int:pid>/fhir')
@permission_required('patients_view')
def patient_fhir_export(pid):
    patient = db.get_patient(pid)
    if not patient:
        return jsonify({'error': 'Not found'}), 404
    labs   = db.get_lab_tests(pid)
    meds   = db.get_prescriptions(pid)
    import uuid
    bundle_id = str(uuid.uuid4())
    entries = []
    # Patient resource
    fhir_patient = {
        'resourceType': 'Patient',
        'id': 'pat-' + str(pid),
        'name': [{'family': patient['last_name'], 'given': [patient['first_name']]}],
        'gender': (patient.get('gender') or '').lower() or 'unknown',
        'birthDate': str(patient['dob']) if patient.get('dob') else None,
        'telecom': [{'system':'phone','value': patient.get('phone')}] if patient.get('phone') else [],
        'address': [{'text': patient.get('address')}] if patient.get('address') else [],
        'identifier': [{'system':'local','value': str(pid)}]
    }
    entries.append({'resource': fhir_patient})
    # Medication resources
    for m in meds:
        entries.append({'resource': {
            'resourceType': 'MedicationStatement',
            'id': 'med-' + str(m['id']),
            'subject': {'reference': 'Patient/pat-' + str(pid)},
            'medicationCodeableConcept': {'text': m.get('drug_name','')},
            'dosage': [{'text': (m.get('dosage') or '') + ' ' + (m.get('frequency') or '')}],
            'status': m.get('status','active')
        }})
    # Lab result resources
    for lt in labs:
        entries.append({'resource': {
            'resourceType': 'Observation',
            'id': 'obs-' + str(lt['id']),
            'subject': {'reference': 'Patient/pat-' + str(pid)},
            'code': {'text': lt.get('test_name','')},
            'valueString': lt.get('result_value'),
            'referenceRange': [{'text': lt.get('reference_range')}] if lt.get('reference_range') else [],
            'interpretation': [{'text': 'abnormal'}] if lt.get('is_abnormal') else [],
            'status': 'final' if lt.get('result_date') else 'registered'
        }})
    bundle = {
        'resourceType': 'Bundle', 'id': bundle_id,
        'type': 'collection',
        'timestamp': datetime.now().isoformat() + 'Z',
        'entry': entries
    }
    db.log_activity(session['staff_id'], 'PHI EXPORT', 'FHIR export — patient #' + str(pid))
    from flask import Response
    return Response(json.dumps(bundle, default=str, indent=2),
                    mimetype='application/fhir+json',
                    headers={'Content-Disposition': 'attachment;filename=patient_' + str(pid) + '_fhir.json'})

# ── Inventory ─────────────────────────────────────────────────────────────────

@app.route('/inventory')
@permission_required('billing')
def inventory():
    items = db.get_inventory()
    return render_template('inventory.html', user=current_user(),
                           practice_name=PRACTICE_NAME, items=items)

@app.route('/inventory/save', methods=['POST'])
@permission_required('billing')
def inventory_save():
    item_id = request.form.get('item_id') or None
    if item_id: item_id = int(item_id)
    data = {'item_name': request.form.get('item_name','').strip(),
            'category':  request.form.get('category',''),
            'quantity':  int(request.form.get('quantity',0) or 0),
            'unit':      request.form.get('unit',''),
            'reorder_level': int(request.form.get('reorder_level',5) or 5),
            'cpt_code':  request.form.get('cpt_code','').strip() or None,
            'unit_cost': request.form.get('unit_cost') or None}
    db.save_inventory_item(data, item_id)
    flash('Inventory updated.', 'success')
    return redirect(url_for('inventory'))

@app.route('/inventory/<int:iid>/use', methods=['POST'])
@login_required
def inventory_use(iid):
    qty = int(request.form.get('qty', 1) or 1)
    db.decrement_inventory(iid, qty)
    return redirect(request.referrer or url_for('inventory'))

# ── Referral Close-the-Loop ────────────────────────────────────────────────────

@app.route('/referrals/<int:ref_id>/outcome', methods=['POST'])
@permission_required('clinical')
def referral_outcome(ref_id):
    outcome         = request.form.get('outcome', 'pending')
    notes           = request.form.get('outcome_notes', '').strip() or None
    appt_confirmed  = request.form.get('appointment_confirmed') == 'on'
    report_received = request.form.get('report_received') == 'on'
    db.update_referral_outcome(ref_id, outcome, notes, appt_confirmed, report_received)
    db.log_activity(session['staff_id'], 'Referral outcome updated',
                    'Ref #' + str(ref_id) + ' — ' + outcome)
    flash('Referral outcome recorded.', 'success')
    return redirect(request.referrer or url_for('dashboard'))

# ── Wait-Time API ─────────────────────────────────────────────────────────────

@app.route('/api/wait-time')
@login_required
def api_wait_time():
    stats = db.get_wait_time_stats()
    return jsonify({'ok': True, 'stats': dict(stats) if stats else {}})


# ════════════════════════════════════════════
#  Research Query Builder + Team Chat
# ════════════════════════════════════════════



# ── Form PDF Downloads & Custom Upload ────────────────────────────────────────

BUILTIN_FORMS = {
    'intake_new_patient': {'title': 'Formulari i Pacientit të Ri',        'icon': 'bi-person-plus-fill',    'color': '#1d4ed8'},
    'consent_general':    {'title': 'Pëlqim i Përgjithshëm',              'icon': 'bi-shield-check',        'color': '#059669'},
    'consent_privacy':    {'title': 'Njoftim mbi Privatësinë',            'icon': 'bi-lock-fill',           'color': '#7c3aed'},
    'consent_telehealth': {'title': 'Pëlqim për Telehealth',              'icon': 'bi-camera-video-fill',   'color': '#0891b2'},
    'consent_photo':      {'title': 'Fotografim / Imazhe Klinike',        'icon': 'bi-camera-fill',         'color': '#d97706'},
    'consent_medication': {'title': 'Pëlqim për Medikamentin',            'icon': 'bi-capsule',             'color': '#dc2626'},
}

@app.route('/forms/pdf/<form_type>')
@login_required
def form_pdf_download(form_type):
    import os
    from flask import send_from_directory
    forms_dir = os.path.join(app.root_path, 'static', 'forms')
    # Check custom uploaded form first
    custom_path = os.path.join(forms_dir, 'custom', secure_filename(form_type) + '.pdf')
    if os.path.exists(custom_path):
        return send_from_directory(os.path.dirname(custom_path), f'{form_type}.pdf', as_attachment=True)
    # Fall back to built-in
    builtin_path = os.path.join(forms_dir, f'{form_type}.pdf')
    if os.path.exists(builtin_path):
        return send_from_directory(forms_dir, f'{form_type}.pdf', as_attachment=True)
    flash('Formulari nuk u gjet.', 'danger')
    return redirect('/forms/library')

@app.route('/forms/library')
@login_required
def forms_library():
    import os
    forms_dir = os.path.join(app.root_path, 'static', 'forms')
    custom_dir = os.path.join(forms_dir, 'custom')
    os.makedirs(custom_dir, exist_ok=True)
    # List custom uploaded forms
    custom_forms = []
    for f in os.listdir(custom_dir):
        if f.endswith('.pdf'):
            name = f.replace('.pdf','').replace('_',' ').title()
            custom_forms.append({'key': f.replace('.pdf',''), 'title': name, 'filename': f})
    return render_template('forms_library.html',
        user=current_user(), practice_name=PRACTICE_NAME,
        builtin_forms=BUILTIN_FORMS,
        custom_forms=custom_forms)

@app.route('/forms/upload', methods=['POST'])
@admin_required
def form_pdf_upload():
    import os
    from werkzeug.utils import secure_filename
    file = request.files.get('pdf_file')
    form_name = request.form.get('form_name','').strip()
    if not file or not file.filename.endswith('.pdf'):
        flash('Ju lutemi ngarkoni një skedar PDF.', 'danger')
        return redirect('/forms/library')
    if not form_name:
        flash('Emri i formularit është i detyrueshëm.', 'danger')
        return redirect('/forms/library')
    safe_name = secure_filename(form_name.lower().replace(' ','_')) + '.pdf'
    custom_dir = os.path.join(app.root_path, 'static', 'forms', 'custom')
    os.makedirs(custom_dir, exist_ok=True)
    file.save(os.path.join(custom_dir, safe_name))
    db.log_activity(session['staff_id'], 'Ngarkoi formular PDF', form_name)
    flash(f'Formulari "{form_name}" u ngarkua me sukses.', 'success')
    return redirect('/forms/library')

@app.route('/forms/delete/<form_key>', methods=['POST'])
@admin_required
def form_pdf_delete(form_key):
    import os
    safe_key = secure_filename(form_key)
    path = os.path.join(app.root_path, 'static', 'forms', 'custom', f'{safe_key}.pdf')
    if os.path.exists(path):
        os.remove(path)
        db.log_activity(session['staff_id'], 'Fshiu formular PDF', form_key)
        flash('Formulari u fshi.', 'success')
    return redirect('/forms/library')

@app.route('/campaigns', methods=['GET','POST'])
@login_required
def campaigns():
    doctor_phone    = None
    doctor_platform = 'whatsapp'
    message         = session.get('campaign_message')
    saved_topic     = session.get('campaign_topic','')
    session_message = session.get('campaign_message')
    results         = None
    filters         = {}

    try:
        s = db.get_settings()
        doctor_phone    = s.get('campaign_doctor_phone','')
        doctor_platform = s.get('campaign_doctor_platform','whatsapp')
    except Exception:
        pass

    if request.method == 'POST':
        action = request.form.get('action','')

        if action == 'save_phone':
            phone    = request.form.get('doctor_phone','').strip()
            platform = request.form.get('doctor_platform','wa')
            try:
                db.save_setting('campaign_doctor_phone', phone)
                db.save_setting('campaign_doctor_platform', platform)
            except Exception:
                pass
            return redirect('/campaigns?step=2')

        elif action == 'generate_message':
            topic = request.form.get('message_topic','').strip()
            session['campaign_topic'] = topic
            if topic:
                try:
                    s = db.get_settings()
                    pname = s.get('practice_name', PRACTICE_NAME)
                    prompt = f"""Shkruaj një mesazh të shkurtër WhatsApp/Viber në Shqip (max 3 fjali) për pacientët e klinikës "{pname}".
Tema: {topic}
Mesazhi duhet të jetë profesional, miqësor dhe të përmbajë ftesë për kontakt."""
                    msg = _gemini(prompt)
                    session['campaign_message'] = msg
                except Exception as e:
                    session['campaign_message'] = f'Mesazh nga {PRACTICE_NAME}: {topic}. Ju ftojmë të kontaktoni klinikën tonë.'
            return redirect('/campaigns?step=2&generated=1')

        elif action == 'save_message':
            msg = request.form.get('message','').strip()
            session['campaign_message'] = msg
            return redirect('/campaigns')

        elif action == 'save_message_and_continue':
            edited = request.form.get('edited_message','').strip()
            if edited:
                session['campaign_message'] = edited
            return redirect('/campaigns?step=3')

        elif action == 'search':
            filters = {
                'age_min':            request.form.get('age_min',''),
                'age_max':            request.form.get('age_max',''),
                'gender':             request.form.get('gender','all'),
                'no_visit_months':    request.form.get('no_visit_months',''),
                'diagnosis_contains': request.form.get('diagnosis_contains',''),
                'icd_contains':       request.form.get('icd_contains',''),
                'has_phone_only':     bool(request.form.get('has_phone_only')),
            }
            try:
                results = db.get_campaign_patients(filters)
                session['campaign_filters'] = filters
            except Exception:
                results = []
            current_step = 4
            searched = True
            if results is None:
                results = []
            return render_template('campaigns.html',
                               user=current_user(),
                               practice_name=PRACTICE_NAME,
                               doctor_phone=doctor_phone,
                               doctor_platform=doctor_platform,
                               message=message,
                               saved_topic=saved_topic,
                               session_message=session_message,
                               results=results,
                               filters=filters,
                               current_step=4,
                               generated=True,
                               searched=True,
                               phone_confirmed=bool(doctor_phone),
                               needing_contact=[],
                               today=__import__('datetime').date.today().isoformat())

    current_step = int(request.args.get('step', 1))
    generated    = request.args.get('generated') == '1'
    searched     = results is not None
    if results is None:
        results = []
    return render_template('campaigns.html',
                           user=current_user(),
                           practice_name=PRACTICE_NAME,
                           doctor_phone=doctor_phone,
                           doctor_platform=doctor_platform,
                           message=message,
                           saved_topic=saved_topic,
                           session_message=session_message,
                           results=results,
                           filters=filters,
                           current_step=current_step,
                           generated=generated,
                           searched=searched,
                           phone_confirmed=bool(doctor_phone),
                           needing_contact=[],
                           today=__import__('datetime').date.today().isoformat())

@app.route('/research-query')
@login_required
def research_query():
    return redirect(url_for('campaigns'))


@app.route('/team-chat')
@login_required
def team_chat():
    messages_list = db.get_team_chat(200)
    staff = db.get_all_staff()
    return render_template('team_chat.html', user=current_user(),
        practice_name=PRACTICE_NAME, messages=messages_list, staff=staff)

@app.route('/team-chat/send', methods=['POST'])
@login_required
def team_chat_send():
    body = request.form.get('body','').strip()
    if body:
        db.send_team_message(session['staff_id'], body)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'ok': True})
    return redirect(url_for('team_chat'))

@app.route('/api/team-chat/messages')
@login_required
def api_team_chat_messages():
    since = request.args.get('since')  # ISO timestamp
    msgs  = db.get_team_chat(200)
    if since:
        from datetime import datetime as dt
        try:
            since_dt = dt.fromisoformat(since)
            msgs = [m for m in msgs if m['created_at'] > since_dt]
        except: pass
    return jsonify({'ok': True, 'messages': [
        {'id': m['id'], 'sender': m['sender_name'], 'role': m['sender_role'],
         'body': m['body'], 'ts': m['created_at'].isoformat()}
        for m in msgs
    ]})


@app.route('/uploads/<path:filename>')
@login_required
def serve_upload(filename):
    # Block path traversal: reject any filename containing directory separators
    import posixpath
    safe_name = os.path.basename(posixpath.normpath('/' + filename))
    if not safe_name or safe_name != filename.split('/')[-1]:
        from flask import abort
        abort(404)
    return send_from_directory(UPLOAD_FOLDER, safe_name)

# ── Medicine Catalog API ──────────────────────────────────────────────────────

@app.route('/api/medicine-search')
@login_required
def api_medicine_search():
    q            = request.args.get('q','').strip()
    in_stock     = request.args.get('in_stock','') == '1'
    country      = request.args.get('country','').strip()
    drug_class   = request.args.get('drug_class','').strip()
    results      = db.search_medicine_catalog(q, in_stock, country, drug_class)
    countries, classes = db.get_medicine_catalog_filters()
    return jsonify({
        'ok': True,
        'results': [dict(r) for r in results],
        'countries': countries,
        'classes': classes,
    })

@app.route('/api/medicine-prescribe', methods=['POST'])
@login_required
def api_medicine_prescribe():
    body       = request.json or {}
    patient_id = body.get('patient_id')
    medicines  = body.get('medicines', [])
    if not patient_id or not medicines:
        return jsonify({'ok': False, 'error': 'Missing data'}), 400
    ids = db.save_patient_medicines(int(patient_id), session['staff_id'], medicines)
    db.log_activity(session['staff_id'], 'Medicines prescribed',
                    f"Patient #{patient_id} — {len(ids)} medicine(s)")
    return jsonify({'ok': True, 'count': len(ids), 'ids': ids})


# ── Security Profile ──────────────────────────────────────────────────────────

@app.route('/security-profile', methods=['GET','POST'])
@login_required
def security_profile():
    staff = db.get_staff_by_id(session['staff_id'])
    if request.method == 'POST' and request.form.get('action') == 'change_password':
        cur_pw  = request.form.get('current_password','')
        new_pw  = request.form.get('new_password','')
        conf_pw = request.form.get('confirm_password','')
        if not bcrypt.check_password_hash(staff['password'], cur_pw):
            flash('Fjalëkalimi aktual është i gabuar.', 'danger')
        elif new_pw != conf_pw:
            flash('Fjalëkalimet e reja nuk përputhen.', 'danger')
        elif len(new_pw) < 12:
            flash('Fjalëkalimi duhet të ketë të paktën 12 karaktere.', 'danger')
        else:
            from security import validate_password_strength
            errors = validate_password_strength(new_pw, staff['email'])
            if errors:
                flash(errors[0], 'danger')
            else:
                hashed = bcrypt.generate_password_hash(new_pw).decode('utf-8')
                db.save_staff({'name': staff['name'], 'email': staff['email'],
                               'role': staff['role'], 'specialty': staff.get('specialty',''),
                               'phone': staff.get('phone',''), 'active': staff['active']},
                              sid=staff['id'])
                conn = db.get_conn(); cur2 = conn.cursor()
                cur2.execute('UPDATE staff SET password=%s WHERE id=%s', (hashed, staff['id']))
                conn.commit(); cur2.close(); conn.close()
                db.log_activity(session['staff_id'], 'Password changed', 'Staff updated own password')
                flash('Fjalëkalimi u ndryshua me sukses.', 'success')
        return redirect('/security-profile')

    has_mfa    = bool(staff and staff.get('mfa_secret'))
    login_ip   = session.get('login_ip', session.get('bound_ip', '—'))
    timeout_min = 60
    return render_template('security_profile.html', user=current_user(),
                           practice_name=PRACTICE_NAME, staff=staff,
                           has_mfa=has_mfa, login_ip=login_ip,
                           timeout_min=timeout_min)



@app.route('/api/mfa-rotation', methods=['POST'])
@admin_required
def api_mfa_rotation():
    try:
        result = sec.rotate_all_mfa_secrets()
        db.log_activity(session['staff_id'], 'MFA Rotation manuale',
                        f"Rotuar: {result.get('rotated',0)}, Anashkaluar: {result.get('skipped',0)}")
        return jsonify({'ok': True, 'rotated': result.get('rotated', 0), 'skipped': result.get('skipped', 0)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

# ── Security Monitoring Dashboard ─────────────────────────────────────────────

@app.route('/security-dashboard')
@admin_required
def security_dashboard():
    try:
        audit = db.get_activity_log(50)
    except Exception:
        audit = []
    events  = []
    summary = {}
    return render_template('security_dashboard.html',
                           user=current_user(),
                           practice_name=PRACTICE_NAME,
                           events=events,
                           summary=summary,
                           audit=audit)

@app.route('/api/security-summary')
@admin_required
def api_security_summary():
    return jsonify(sec.get_security_summary())

@app.route('/api/security-events')
@admin_required
def api_security_events():
    hours  = int(request.args.get('hours', 24))
    events = sec.get_security_events(hours=hours, limit=200)
    # Convert datetime objects for JSON
    for e in events:
        if e.get('created_at'):
            e['created_at'] = e['created_at'].isoformat()
    return jsonify({'ok': True, 'events': events})


# ── CSP Violation Report Endpoint ─────────────────────────────────────────────

@app.route('/api/csp-report', methods=['POST'])
def csp_report():
    """Receives CSP violation reports from browsers. Set CSP_REPORT_URI=/api/csp-report"""
    try:
        report = request.get_json(force=True, silent=True) or {}
        violation = report.get('csp-report', report)
        db.log_activity(
            session.get('staff_id'),
            'CSP_VIOLATION',
            f"blocked={violation.get('blocked-uri','?')} "
            f"directive={violation.get('violated-directive','?')} "
            f"doc={violation.get('document-uri','?')[:80]}"
        )
    except Exception:
        pass
    return '', 204   # No Content


# ── Recalls (top-level — shows all pending recalls across all patients) ────────

# ── Cookie / Fingerprint Consent ──────────────────────────────────────────────

@app.route('/consent/fingerprint', methods=['POST'])
@login_required
def set_fp_consent():
    """Set or revoke browser fingerprint consent (GDPR)."""
    granted  = request.json.get('granted', False) if request.is_json else                request.form.get('granted') == '1'
    response = jsonify({'ok': True, 'granted': granted})
    sec.set_fingerprint_consent(response, granted)
    db.log_activity(session['staff_id'], 'FINGERPRINT_CONSENT',
                    f"granted={granted}")
    return response


# ── Password Reset ─────────────────────────────────────────────────────────────

@app.route('/reset-password', methods=['GET', 'POST'])
def reset_password_request():
    """
    Step 1 — staff enters email.
    A one-time token is stored (hashed) and would be emailed.
    Token is valid for 30 minutes.
    """
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        staff = db.get_staff_by_email(email)
        # Always show same message — prevents email enumeration
        flash('If that email exists, a reset link has been sent.', 'info')
        if staff and staff.get('active'):
            import hmac as _hmac
            token    = secrets.token_urlsafe(32)
            expires  = (datetime.now() + timedelta(minutes=30)).isoformat()
            tok_hash = _hmac.new(
                app.secret_key.encode(), token.encode(), 'sha256'
            ).hexdigest()
            # Store hash + expiry (in-memory for simplicity; use DB in production)
            _pw_reset_tokens[tok_hash] = {'staff_id': staff['id'], 'expires': expires}
            db.log_activity(staff['id'], 'PASSWORD_RESET_REQUESTED',
                            f"email={email}")
            # In production: email the token link to the user
            # For now: print to server log (dev only)
            reset_url = f"http://localhost:5000/reset-password/{token}"
            print(f"[DEV ONLY] Password reset link: {reset_url}")
        return redirect(url_for('login'))
    return render_template('reset_password.html', practice_name=PRACTICE_NAME)


# In-memory token store — replace with DB table in production
_pw_reset_tokens: dict = {}


@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password_confirm(token):
    """Step 2 — staff clicks link, enters new password."""
    import hmac as _hmac
    tok_hash = _hmac.new(
        app.secret_key.encode(), token.encode(), 'sha256'
    ).hexdigest()

    entry = _pw_reset_tokens.get(tok_hash)
    if not entry or datetime.fromisoformat(entry['expires']) < datetime.now():
        flash('Reset link is invalid or has expired.', 'danger')
        return redirect(url_for('login'))

    if request.method == 'POST':
        new_pw   = request.form.get('new_password', '')
        conf_pw  = request.form.get('confirm_password', '')
        errors   = sec.validate_password_with_hibp(new_pw)
        if errors:
            flash(errors[0], 'danger')
        elif new_pw != conf_pw:
            flash('Passwords do not match.', 'danger')
        else:
            hashed = bcrypt.generate_password_hash(new_pw).decode('utf-8')
            conn   = db.get_conn(); cur = conn.cursor()
            cur.execute('UPDATE staff SET password=%s WHERE id=%s',
                        (hashed, entry['staff_id']))
            conn.commit(); cur.close(); conn.close()
            _pw_reset_tokens.pop(tok_hash, None)
            db.log_activity(entry['staff_id'], 'PASSWORD_RESET_COMPLETED', '')
            flash('Password reset successfully. Please log in.', 'success')
            return redirect(url_for('login'))

    return render_template('reset_password_confirm.html',
                           practice_name=PRACTICE_NAME, token=token)


# ── Patient Report — Print & Email ────────────────────────────────────────────

@app.route('/patients/<int:pid>/report')
@login_required
def patient_report(pid):
    """Printable patient report with visit notes and prescriptions."""
    patient     = db.get_patient(pid)
    if not patient:
        return redirect('/patients')
    visits      = db.get_visit_notes_for_patient(pid)
    prescriptions = db.get_prescriptions(pid)
    staff_list  = db.get_all_staff()
    from datetime import date as _date
    age = None
    if patient['dob']:
        age = (_date.today() - patient['dob']).days // 365
    return render_template('patient_report.html',
                           user=current_user(),
                           practice_name=PRACTICE_NAME,
                           practice_settings=db.get_settings(),
                           patient=patient,
                           visits=visits,
                           prescriptions=prescriptions,
                           staff_list=staff_list,
                           age=age,
                           today=_date.today())

@app.route('/patients/<int:pid>/report/email', methods=['POST'])
@login_required
def patient_report_email(pid):
    """Send patient report by email."""
    patient = db.get_patient(pid)
    if not patient:
        return jsonify({'ok': False, 'error': 'Patient not found'}), 404

    to_email = request.json.get('email', '').strip() if request.is_json else request.form.get('email','').strip()
    if not to_email:
        to_email = patient.get('email','')
    if not to_email:
        return jsonify({'ok': False, 'error': 'No email address available'}), 400

    visits        = db.get_visit_notes_for_patient(pid)
    prescriptions = db.get_prescriptions(pid)
    from datetime import date as _date
    age = (_date.today() - patient['dob']).days // 365 if patient['dob'] else None

    # Render report as HTML for email
    html_body = render_template('patient_report.html',
                                user=current_user(),
                                practice_name=PRACTICE_NAME,
                                practice_settings=db.get_settings(),
                                patient=patient,
                                visits=visits,
                                prescriptions=prescriptions,
                                staff_list=db.get_all_staff(),
                                age=age,
                                today=_date.today(),
                                email_mode=True)

    try:
        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText

        if not SMTP_USER or not SMTP_PASSWORD:
            return jsonify({'ok': False, 'error': 'Email not configured. Add SMTP settings to config.py'}), 500

        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"Raporti mjekësor — {patient['first_name']} {patient['last_name']}"
        msg['From']    = SMTP_FROM or SMTP_USER
        msg['To']      = to_email
        msg.attach(MIMEText(html_body, 'html'))

        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.send_message(msg)

        db.log_activity(session['staff_id'], 'Sent patient report by email',
                        f"{patient['first_name']} {patient['last_name']} → {to_email}")
        return jsonify({'ok': True, 'message': f'Report sent to {to_email}'})

    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ── Global Forms Hub ──────────────────────────────────────────────────────────

@app.route('/forms')
@login_required
def forms_hub():
    status  = request.args.get('status', '')
    search  = request.args.get('q', '').strip()
    all_forms = db.get_all_forms(status=status or None)
    if search:
        sl = search.lower()
        all_forms = [f for f in all_forms if sl in (f.get('first_name') or '').lower()
                     or sl in (f.get('last_name') or '').lower()
                     or sl in (f.get('form_title') or '').lower()]
    patients_list = db.get_all_patients()
    return render_template('forms_hub.html', user=current_user(),
                           practice_name=PRACTICE_NAME,
                           forms=all_forms,
                           form_templates=db.FORM_TEMPLATES,
                           patients_list=patients_list,
                           status_filter=status,
                           search=search,
                           today=date.today())

@app.route('/forms/send', methods=['POST'])
@login_required
def forms_hub_send():
    """Send a form to a patient from the global hub."""
    pid       = request.form.get('patient_id')
    form_type = request.form.get('form_type', 'consent_general')
    template  = db.FORM_TEMPLATES.get(form_type, {})
    if not pid:
        flash('Please select a patient.', 'danger')
        return redirect('/forms')
    data = {
        'patient_id':   int(pid),
        'form_type':    form_type,
        'form_title':   request.form.get('form_title') or template.get('title', 'Form'),
        'form_content': template.get('content', ''),
    }
    fid = db.save_patient_form(data, created_by=session['staff_id'])
    db.log_activity(session['staff_id'], 'Sent patient form', f"Patient #{pid} — {data['form_title']}")
    flash('Form sent. Copy the link and share it with the patient.', 'success')
    return redirect(f'/forms?highlight={fid}')


# ── Practice Settings ─────────────────────────────────────────────────────────

@app.route('/settings', methods=['GET', 'POST'])
@admin_required
def practice_settings():
    if request.method == 'POST':
        data = {}
        # Text fields
        for key in ['practice_name', 'primary_color', 'accent_color', 'sidebar_color', 'font_family', 'country_code']:
            val = request.form.get(key, '').strip()
            if val:
                data[key] = val

        # Logo upload
        logo_file = request.files.get('logo_file')
        if logo_file and logo_file.filename:
            ext = logo_file.filename.rsplit('.', 1)[-1].lower()
            if ext in {'png', 'jpg', 'jpeg', 'gif', 'svg', 'webp'}:
                import uuid
                logo_name = f"logo_{uuid.uuid4().hex[:8]}.{ext}"
                logo_path  = os.path.join(UPLOAD_FOLDER, logo_name)
                logo_file.save(logo_path)
                data['logo_filename'] = logo_name

        # Remove logo
        if request.form.get('remove_logo'):
            data['logo_filename'] = ''

        if data:
            db.save_settings(data)
            # Update practice name in config-like variable
            global PRACTICE_NAME
            if 'practice_name' in data:
                PRACTICE_NAME = data['practice_name']

        db.log_activity(session['staff_id'], 'Updated practice settings', '')
        flash('Cilësimet u ruajtën.', 'success')
        return redirect('/settings')

    settings = db.get_settings()
    return render_template('settings.html', user=current_user(),
                           practice_name=PRACTICE_NAME,
                           settings=settings)


# ── Run ────────────────────────────────────────────────────────────────────────





# ── Patient Forms ─────────────────────────────────────────────────────────────

@app.route('/patients/<int:pid>/forms')
@login_required
def patient_forms(pid):
    patient = db.get_patient(pid)
    if not patient:
        return redirect('/patients')
    forms = db.get_patient_forms(pid)
    return render_template('patient_forms.html', user=current_user(),
                           practice_name=PRACTICE_NAME, patient=patient,
                           forms=forms, form_templates=db.FORM_TEMPLATES)

@app.route('/patients/<int:pid>/forms/send', methods=['POST'])
@login_required
def patient_form_send(pid):
    form_type = request.form.get('form_type', 'consent_general')
    template  = db.FORM_TEMPLATES.get(form_type, {})
    data = {
        'patient_id':   pid,
        'form_type':    form_type,
        'form_title':   request.form.get('form_title', template.get('title', 'Form')),
        'form_content': request.form.get('form_content', template.get('content', '')),
    }
    fid = db.save_patient_form(data, created_by=session['staff_id'])
    db.log_activity(session['staff_id'], 'Sent patient form', f"Patient ID {pid} — {data['form_title']}")
    flash('Form sent to patient.', 'success')
    return redirect(f'/patients/{pid}/forms')

@app.route('/forms/<int:form_id>/sign', methods=['GET', 'POST'])
def form_sign(form_id):
    """Patient-facing form signing page — no login required."""
    form = db.get_patient_form(form_id)
    if not form:
        return 'Form not found.', 404
    if form['status'] == 'signed':
        return render_template('form_signed_confirm.html', form=form,
                               practice_name=PRACTICE_NAME)
    if request.method == 'POST':
        signed_by = request.form.get('signed_by_name', '').strip()
        sig_data  = request.form.get('signature_data', '')
        ip        = request.remote_addr
        if signed_by and sig_data:
            db.sign_patient_form(form_id, signed_by, sig_data, ip)
            db.log_activity(None, 'Patient form signed',
                            f"Form #{form_id} signed by {signed_by}")
            return render_template('form_signed_confirm.html', form=form,
                                   practice_name=PRACTICE_NAME)
    return render_template('form_sign.html', form=form, practice_name=PRACTICE_NAME)

@app.route('/patients/<int:pid>/forms/<int:form_id>/delete', methods=['POST'])
@login_required
def patient_form_delete(pid, form_id):
    db.delete_patient_form(form_id)
    db.log_activity(session['staff_id'], 'Deleted patient form', f"Form #{form_id}")
    flash('Form deleted.', 'success')
    return redirect(f'/patients/{pid}/forms')


if __name__ == '__main__':
    try:
        db.setup()
        db.setup_settings()
        db.setup_medicine_catalog()
        db.setup_canary_accounts()
        sec._ensure_lockout_table()
        sec.init_scheduler(app)
        db.setup_tasks()
        db.setup_advanced()
        db.setup_team_chat()
        print("\n" + "=" * 55)
        print(f"  {PRACTICE_NAME} — Ready!")
        print("  http://localhost:5000")
        print("  Login: admin@practice.local / admin123")
        print("=" * 55 + "\n")
    except Exception as e:
        print(f"\n[ERROR] Database connection failed: {e}")
        print("Check config.py and ensure PostgreSQL is running.\n")
        exit(1)
    try:
        from waitress import serve
        print('  Server: Waitress (production WSGI)')
        serve(app, host='0.0.0.0', port=5000, threads=8)
    except ImportError:
        print('  Server: Flask dev — run: pip install waitress')
        app.run(host='0.0.0.0', port=5000, debug=False)
